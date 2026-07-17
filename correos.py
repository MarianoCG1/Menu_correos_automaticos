"""
correos_etv_mejorado.py
=======================
Generador de Correos ETV con Carga Flexible de Excel
- Detecta automáticamente estructura del Excel
- Valida columnas críticas
- Preview de datos cargados
- Log detallado de errores
- Interfaz mejorada
"""

import os
import sys
import json
import socket
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
import pandas as pd
from datetime import datetime, date, timedelta
import re
import logging
from tkcalendar import DateEntry

# ─── Configuración ────────────────────────────────────

PORTAL_URL = None
USER_NAME = os.environ.get('USERNAME', 'desconocido')
PC_NAME = socket.gethostname()

CAMPAIGN_CODES = {
    'EF': 'ENERO – FEBRERO', 'MA': 'MARZO – ABRIL',
    'MJ': 'MAYO – JUNIO', 'JA': 'JULIO – AGOSTO',
    'SO': 'SEPTIEMBRE – OCTUBRE', 'ND': 'NOVIEMBRE – DICIEMBRE',
}

# Columnas críticas que DEBEN existir (con posibles variaciones)
REQUIRED_COLUMNS = {
    'Agencia': ['Agencia', 'AGENCIA', 'Ag.', 'Agency', 'BN Agencia', 'Agencia BN', 'Sede', 'AA', 'AG'],
    'Departamento': ['Departamento', 'DEPARTAMENTO', 'Dpto', 'Depto', 'DD', 'DEP'],
    'Provincia': ['Provincia', 'PROVINCIA', 'Prov', 'PP', 'PROV'],
    'Distrito': ['Distrito', 'DISTRITO', 'Dist', 'dd', 'DIST'],
    'Punto_de_Pago': ['Punto de Pago', 'Punto_de_Pago', 'PUNTO DE PAGO', 'Pto Pago', 'pp', 'PPP'],
    'Fecha_Pago': ['Fecha Pago', 'Fecha_Pago', 'FECHA PAGO', 'Fec. Pago', 'Fecha\nPropuesta', 'Fecha Propuesta', 'Fecha Propuesta '],
    'P65_Usuarios': ['P65 Usuarios', 'P65_Usuarios', 'N° Usuarios P65', 'Usuarios P65', 'N° Usuarios', 'U_P65'],
    'P65_Importe': ['P65 Importe', 'P65_Importe', 'Importe P65', 'Imp. P65', 'Importe', 'I_P65'],
}

# ─── Logger Setup ─────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('correos_etv.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)


# ─── Utilidades de Carga de Excel ────────────────────

class ExcelLoader:
    """Carga inteligente de archivos Excel con estructura variable."""
    
    @staticmethod
    def detect_header_row(df, max_rows=10):
        """Detecta qué fila contiene los encabezados más útiles."""
        best_row = 0
        best_score = 0
        
        for i in range(min(max_rows, len(df))):
            row = df.iloc[i].astype(str)
            
            # Contar keywords típicos de encabezados
            keywords = ['agencia', 'departamento', 'provincia', 'fecha', 'usuarios', 'distrito', 'importe']
            keyword_matches = sum(1 for val in row if any(kw in str(val).lower() for kw in keywords))
            
            # Contar columnas NO-Unnamed (más específicas)
            non_unnamed = sum(1 for val in row if 'unnamed' not in str(val).lower() and str(val).strip())
            
            # Penalizar filas con muchos NaN
            nan_count = row.isna().sum()
            
            # Score: favorece keywords + columnas específicas - NaNs
            score = (keyword_matches * 10) + non_unnamed - (nan_count * 2)
            
            logging.info(f"Fila {i}: keywords={keyword_matches}, non_unnamed={non_unnamed}, nans={nan_count}, score={score}")
            
            if score > best_score:
                best_score = score
                best_row = i
        
        logging.info(f"Header detectado en fila {best_row} con score {best_score}")
        return best_row
    
    @staticmethod
    def normalize_column_name(col_name):
        """Normaliza nombres de columnas para comparación."""
        return str(col_name).strip().upper().replace(' ', '_').replace('.', '').replace('°', '')
    
    @staticmethod
    def find_column(df, possible_names, ignore_cols=None):
        """Encuentra una columna por nombres posibles (case-insensitive pero preservando original)."""
        if ignore_cols is None:
            ignore_cols = set()
            
        # Para cada columna del DataFrame, guardar su nombre original y una versión limpia para comparación
        df_cols_map = {}
        for col in df.columns:
            if col in ignore_cols:
                continue
            # Versión normalizada solo para comparación
            clean = str(col).strip().replace(' ', '').replace('.', '').replace('°', '').replace('\n', '').lower()
            if clean not in df_cols_map:
                df_cols_map[clean] = col
        
        for name in possible_names:
            # Normalizar el nombre buscado
            clean_search = str(name).strip().replace(' ', '').replace('.', '').replace('°', '').replace('\n', '').lower()
            if clean_search in df_cols_map:
                return df_cols_map[clean_search]
        return None
    
    @staticmethod
    def load_flexible(path):
        """Carga Excel con detección automática de estructura y hoja."""
        try:
            # 1. Detectar Hoja Correcta
            xls = pd.ExcelFile(path)
            sheet_names = xls.sheet_names
            best_sheet = sheet_names[0]
            for s in sheet_names:
                if any(kw in s.upper() for kw in ['CRONOGRAMA', 'ETV', 'PAGOS', 'PROGRAMA']):
                    best_sheet = s
                    break
            logging.info(f"Usando hoja: {best_sheet}")

            # 2. Cargar sin header para detectar
            df_raw = pd.read_excel(path, sheet_name=best_sheet, header=None)
            header_row = ExcelLoader.detect_header_row(df_raw, max_rows=30)
            
            # 3. Recargar con header correcto
            df = pd.read_excel(path, sheet_name=best_sheet, header=header_row)
            
            # 4. Mapear columnas requeridas
            column_mapping = {}
            missing_columns = []
            matched_cols = set()
            
            for standard_name, possible_names in REQUIRED_COLUMNS.items():
                found_col = ExcelLoader.find_column(df, possible_names, ignore_cols=matched_cols)
                if found_col:
                    column_mapping[found_col] = standard_name
                    matched_cols.add(found_col)
                else:
                    missing_columns.append(standard_name)
            
            # 5. Renombrar y Fallback para Agencia
            df.rename(columns=column_mapping, inplace=True)
            
            if 'Agencia' not in df.columns:
                # Búsqueda manual si fallaron los alias
                potential = [c for c in df.columns if 'agencia' in str(c).lower() and 'Unnamed' not in str(c) and c not in matched_cols]
                if potential:
                    logging.info(f"Fallback detectó Agencia en columna: {potential[0]}")
                    df.rename(columns={potential[0]: 'Agencia'}, inplace=True)
                    matched_cols.add(potential[0])
                    if 'Agencia' in missing_columns: missing_columns.remove('Agencia')
            
            if missing_columns:
                logging.warning(f"Columnas faltantes: {missing_columns}")
            
            # 6. Detectar columnas opcionales (Juntos, Contigo, ETV, etc.)
            optional_mappings = {
                'Juntos_Usuarios': ['Juntos Usuarios', 'N° Usuarios Juntos', 'Usuarios Juntos', 'N° Usuarios.1', 'U_JUNTOS'],
                'Juntos_Importe': ['Juntos Importe', 'Importe Juntos', 'Imp. Juntos', 'Importe.1', 'I_JUNTOS'],
                'Contigo_Usuarios': ['Contigo Usuarios', 'N° Usuarios Contigo', 'Usuarios Contigo', 'N° Usuarios.2', 'U_CONTIGO'],
                'Contigo_Importe': ['Contigo Importe', 'Importe Contigo', 'Imp. Contigo', 'Importe.2', 'I_CONTIGO'],
                'ETV': ['ETV', 'E.T.V.', 'Etv', 'EMPRESA'],
                'Hora_Pago': ['Hora Pago', 'Hora_Pago', 'Hr. Pago', 'Hora', 'HR'],
                'Fecha_Entrega': ['Fecha Entrega', 'Fecha_Entrega', 'Fec. Entrega', 'FECHA', 'FEC_ENT', 'Fecha'],
                'Hora_Entrega': ['Hora Entrega', 'Hora_Entrega', 'Hr. Entrega', 'Hora.1', 'HR_ENT'],
                'Cod_Pueblo': ['Cod Pueblo', 'Código Pueblo', 'Cod. Pueblo', 'cp', 'C_PUEBLO'],
                'Cod_Pto_Pago': ['Cod Pto Pago', 'Código Pto Pago', 'Cod. Pto. Pago', 'Cod. Pto Pago', 'cpp', 'C_PTO'],
            }
            
            for standard_name, possible_names in optional_mappings.items():
                found_col = ExcelLoader.find_column(df, possible_names, ignore_cols=matched_cols)
                if found_col and found_col != standard_name:
                    df.rename(columns={found_col: standard_name}, inplace=True)
                    matched_cols.add(found_col)
            
            # --- Lógica de Fecha Inteligente ---
            # Identificar todas las columnas que parecen fechas para encontrar la mínima
            all_date_cols = []
            
            # Buscar columnas estándar
            for col in ['Fecha_Pago', 'Fecha_Entrega']:
                if col in df.columns:
                    all_date_cols.append(col)
            
            # Buscar columnas adicionales que huelan a fecha si no tenemos suficientes
            if len(all_date_cols) < 2:
                for col in df.columns:
                    if col not in all_date_cols and any(kw in str(col).upper() for kw in ['FECHA', 'FEC', 'PROPUESTA']):
                        # Evitar columnas 'Unnamed' o 'Hora'
                        if 'UNNAMED' not in str(col).upper() and 'HORA' not in str(col).upper():
                            all_date_cols.append(col)
            
            logging.info(f"Columnas de fecha identificadas para cálculo: {all_date_cols}")
            
            # Crear columna de fecha objetivo calculada
            def get_min_date(row):
                dates = []
                for col in all_date_cols:
                    val = row.get(col)
                    if pd.notna(val):
                        try:
                            d = pd.to_datetime(val, errors='coerce', dayfirst=True)
                            if pd.notna(d):
                                dates.append(d)
                        except:
                            pass
                return min(dates).date() if dates else None

            df['Target_Date_Calc'] = df.apply(get_min_date, axis=1)
            
            # 7. Limpiar datos (Solo si encontramos Agencia)
            if 'Agencia' in df.columns:
                df = df.dropna(subset=['Agencia'], how='all')
                df = df[df['Agencia'].astype(str).str.strip().str.lower() != 'nan']
                df = df[df['Agencia'].astype(str).str.strip() != '']
            
            logging.info(f"Excel listo: {len(df)} filas detectadas.")
            if 'Target_Date_Calc' in df.columns and not df['Target_Date_Calc'].isna().all():
                logging.info(f"Fechas detectadas en rango calc: {df['Target_Date_Calc'].min()} a {df['Target_Date_Calc'].max()}")
                # Depuración resumida
                stats = df['Target_Date_Calc'].value_counts().sort_index().head(5)
                logging.info(f"Días próximos detectados:\n{stats}")
            return df, missing_columns
            
        except Exception as e:
            logging.error(f"Error crítico en carga: {e}")
            raise


# ─── Aplicación Principal ─────────────────────────────

class CorreosETVApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Correos ETV — Banco de la Nación (Versión Mejorada)")
        self.root.resizable(True, True)
        self.root.configure(bg='#f4f7f9')

        # Centrar ventana en pantalla
        w, h = 950, 680
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

        # State
        self.df_cronograma = None
        self.df_agencias = None
        self.campaign = ''
        self.campaign_code = ''
        self.cronograma_path = ''
        self.backup_network_path = ''   # Ruta fija de red para backups (configurable)
        self.missing_columns = []
        self.error_log = []
        self.selected_date = date.today() + timedelta(days=1)

        # Build UI FIRST so widgets exist for logging/loading
        self._build_ui()
        
        # Then load data
        self._load_agencias()
        self._load_config()   # cargar campana persistida

        # Cargar icono del programa (User's custom icon)
        try:
            icon_p = self._resource_path('icono_programa.png')
            if os.path.exists(icon_p):
                self.icon_img = tk.PhotoImage(file=icon_p)
                self.root.iconphoto(False, self.icon_img)
        except Exception as e:
            print(f"Aviso: No se pudo cargar el icono del aplicativo: {e}")

    def _load_agencias(self):
        path = self._find_file('agencias_correos.xlsx')
        if path:
            try:
                df = pd.read_excel(path)

                # normalizar encabezados para evitar KeyError más adelante
                # buscamos al menos la columna de agencia y opcionales
                col_map = {}
                ag_cols = ['Agencia', 'AGENCIA', 'AGENCIAS', 'Agencias', 'Ag.', 'Agency', 'BN Agencia', 'Agencia BN', 'Sede', 'AA', 'AG']
                correo_cols = ['Correo', 'Email', 'E-mail', 'Mail', 'E-mail BN', 'E_MAIL', 'JEFE DE PAGO', 'Jefe de Pago', 'Jefe De Pago']
                cc_cols = ['CC', 'Copia', 'C.C.', 'Cco', 'CCO', 'ADMINISTRADOR', 'Administrador']

                # función auxiliar rápida para buscar en df.columns
                def find_col(possibles):
                    for c in df.columns:
                        clean = str(c).strip().lower()
                        for p in possibles:
                            if clean == str(p).strip().lower():
                                return c
                    return None

                found = find_col(ag_cols)
                if found:
                    col_map[found] = 'Agencia'
                found = find_col(correo_cols)
                if found:
                    col_map[found] = 'Correo'
                found = find_col(cc_cols)
                if found:
                    col_map[found] = 'CC'

                if col_map:
                    df.rename(columns=col_map, inplace=True)
                    self._log(f"   [Debug] Columnas renombradas: {col_map}")

                self.df_agencias = df
                self._log(f"📋 Agencias cargadas: {len(self.df_agencias)} registros desde {os.path.basename(path)}")
                self._log(f"   [Debug] Columnas finales: {list(df.columns)}")

                # advertir si no existe la columna esperada
                if 'Agencia' not in df.columns:
                    logging.warning("El archivo de agencias no contiene una columna 'Agencia' reconocida.")
            except Exception as e:
                self._log(f"⚠ No se pudo cargar agencias: {e}")
                self.df_agencias = None
        else:
            self._log("⚠ No se encontró archivo de correos (agencias_correos.xlsx)")
            self.df_agencias = None


    def _load_config(self):
        """Carga configuración persistida (ruta del Excel de campaña y automatización)."""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'etv_config.json')
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                
                # Cargar ruta del cronograma
                saved_path = cfg.get('cronograma_path', '')
                if saved_path and os.path.exists(saved_path):
                    self._load_cronograma_from_path(saved_path)
                
                # Cargar ruta fija de backup en red
                bnp = cfg.get('backup_network_path', '')
                self.backup_network_path = bnp
                if hasattr(self, 'backup_path_var'):
                    self.backup_path_var.set(bnp)
                    self._update_backup_status()
                
                # Cargar configs de automatización
                auto_enabled = cfg.get('auto_enabled', False)
                self.auto_enabled_var.set(auto_enabled)
                
                auto_hour = cfg.get('auto_hour', '08:00')
                self.auto_hour_var.set(auto_hour)
                
                auto_send = cfg.get('auto_send', True)
                self.auto_send_var.set(auto_send)
                
                # Si está activada, arrancar el scheduler
                if auto_enabled:
                    self._start_auto_scheduler()
                    self._update_auto_status_ui(True)
        except Exception as e:
            logging.warning(f"No se pudo cargar config guardada: {e}")

    def _save_config(self):
        """Guarda la ruta del Excel de campaña y la configuración de automatización."""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'etv_config.json')
        try:
            cfg = {
                'cronograma_path': self.cronograma_path,
                'backup_network_path': self.backup_network_path,
                'auto_enabled': self.auto_enabled_var.get() if hasattr(self, 'auto_enabled_var') else False,
                'auto_hour': self.auto_hour_var.get().strip() if hasattr(self, 'auto_hour_var') else '08:00',
                'auto_send': self.auto_send_var.get() if hasattr(self, 'auto_send_var') else True
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.warning(f"No se pudo guardar config: {e}")

    def _resource_path(self, relative_path):
        """ Obtiene ruta absoluta para recursos, compatible con PyInstaller """
        try:
            # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        # Primero intentar en la carpeta de recursos de PyInstaller
        res_p = os.path.join(base_path, relative_path)
        if os.path.exists(res_p):
            return res_p
        return relative_path

    def _find_file(self, filename):
        """Busca archivo en varias ubicaciones posibles."""
        # Directorio del script (.py) o del interior del .exe (PyInstaller _MEIPASS)
        try:
            base_dir = sys._MEIPASS
        except AttributeError:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Directorio del ejecutable (.exe) o del intérprete python
        exe_dir = os.path.dirname(sys.executable)
        
        # Directorio de trabajo actual (desde donde se ejecutó el programa)
        cwd = os.getcwd()
        
        candidates = [
            # Buscar en el directorio del script/exe
            os.path.join(base_dir, filename),
            os.path.join(base_dir, 'data', filename),
            os.path.join(base_dir, 'docs', filename),
            # Buscar en el directorio de trabajo actual
            os.path.join(cwd, filename),
            os.path.join(cwd, 'data', filename),
            os.path.join(cwd, 'docs', filename),
            # Buscar junto al ejecutable
            os.path.join(exe_dir, filename),
            os.path.join(exe_dir, 'data', filename),
            os.path.join(exe_dir, 'docs', filename),
            # Buscar un nivel arriba
            os.path.join(base_dir, '..', filename),
            os.path.join(base_dir, '..', 'data', filename),
            os.path.join(cwd, '..', filename),
            os.path.join(cwd, '..', 'data', filename),
            # Ruta directa
            filename,
        ]
        
        for c in candidates:
            if os.path.exists(c):
                return os.path.abspath(c)
        
        # Si no se encontró, loguear las rutas buscadas para depuración
        self._log(f"   [Debug] No se encontró '{filename}'. Rutas buscadas:")
        self._log(f"      base_dir: {base_dir}")
        self._log(f"      exe_dir:  {exe_dir}")
        self._log(f"      cwd:      {cwd}")
        return None


    def _build_ui(self):
        # ─── Estilos y Sidebar ───
        self.sidebar = tk.Frame(self.root, bg='#1a1c1e', width=160)
        self.sidebar.pack(side='left', fill='y')
        self.sidebar.pack_propagate(False)

        # Logo en Sidebar
        logo_frame = tk.Frame(self.sidebar, bg='#1a1c1e', pady=30)
        logo_frame.pack(fill='x')
        
        logo_path = self._find_file('logo_bn.png')
        if logo_path:
            try:
                # Intenta usar Pillow para redimensionado suave si está instalado
                try:
                    from PIL import Image, ImageTk
                    img = Image.open(logo_path)
                    # El sidebar tiene 160px, usamos 140px como ancho máximo
                    max_width = 140
                    ratio = max_width / img.width
                    new_size = (int(img.width * ratio), int(img.height * ratio))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                    self.logo_img = ImageTk.PhotoImage(img)
                    tk.Label(logo_frame, image=self.logo_img, bg='#1a1c1e').pack()
                except ImportError:
                    # Fallback a PhotoImage puro con subsample
                    raw_img = tk.PhotoImage(file=logo_path)
                    width = raw_img.width()
                    factor = max(1, width // 140)
                    self.logo_img = raw_img.subsample(factor, factor)
                    tk.Label(logo_frame, image=self.logo_img, bg='#1a1c1e').pack()
            except Exception as e:
                logging.warning(f"Error cargando logo: {e}")
                tk.Label(logo_frame, text="BN", bg='#C8102E', fg='white', font=('Segoe UI', 18, 'bold'), width=3, pady=5).pack()
        else:
            tk.Label(logo_frame, text="BN", bg='#C8102E', fg='white', font=('Segoe UI', 18, 'bold'), width=3, pady=5).pack()

        tk.Label(
            logo_frame, text="Portal Automatización", bg='#1a1c1e', fg='white',
            font=('Segoe UI', 9, 'bold'), pady=10
        ).pack()

        # Botones de Sidebar (Navegación)
        self.nav_btns = {}
        nav_info = [
            ('config', '⚙️  Configuración'),
            ('preview', '👁️  Vista Previa'),
            ('log', '📋  Log de Procesos')
        ]
        
        for view, text in nav_info:
            btn = tk.Button(
                self.sidebar, text=text, bg='#1a1c1e', fg='#b2bec3',
                font=('Segoe UI', 10), relief='flat', anchor='w', padx=25, pady=12,
                cursor='hand2', activebackground='#C8102E', activeforeground='white',
                command=lambda v=view: self._show_view(v)
            )
            btn.pack(fill='x')
            self.nav_btns[view] = btn

        # Footer del Sidebar
        sidebar_footer = tk.Frame(self.sidebar, bg='#1a1c1e', pady=20)
        sidebar_footer.pack(side='bottom', fill='x')
        tk.Label(
            sidebar_footer, text=f"v8.1 Standalone", bg='#1a1c1e', fg='#636e72',
            font=('Segoe UI', 8)
        ).pack()

        # ─── Main Content Area ───
        self.main_container = tk.Frame(self.root, bg='#f4f7f9')
        self.main_container.pack(side='right', fill='both', expand=True)

        self.views = {}
        for view in ['config', 'preview', 'log']:
            canvas = tk.Canvas(self.main_container, bg='#f4f7f9', highlightthickness=0)
            scrollbar = ttk.Scrollbar(self.main_container, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='#f4f7f9', padx=20, pady=20)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=800)
            canvas.configure(yscrollcommand=scrollbar.set)
            
            self.views[view] = {'frame': scrollable_frame, 'canvas': canvas, 'scrollbar': scrollbar}
        
        self.active_view = 'config'
        # Bind MouseWheel once: scroll whichever canvas is currently visible
        self.root.bind_all("<MouseWheel>", self._on_main_mousewheel)
            
        self._build_config_tab(self.views['config']['frame'])
        self._build_preview_tab(self.views['preview']['frame'])
        self._build_log_tab(self.views['log']['frame'])
        
        self._show_view('config')

    def _on_main_mousewheel(self, event):
        """Global handler for mouse scroll in main window."""
        cv = self.views.get(self.active_view, {}).get('canvas')
        if cv and cv.winfo_exists():
            try:
                cv.yview_scroll(int(-1*(event.delta/120)), "units")
            except Exception:
                pass

    def _show_view(self, view_name):
        self.active_view = view_name
        for name, components in self.views.items():
            if name == view_name:
                components['canvas'].pack(side="left", fill="both", expand=True)
                components['scrollbar'].pack(side="right", fill="y")
                self.nav_btns[name].config(bg='#C8102E', fg='white')
            else:
                components['canvas'].pack_forget()
                components['scrollbar'].pack_forget()
                self.nav_btns[name].config(bg='#1a1c1e', fg='#b2bec3')

    def _build_config_tab(self, parent):
        """Pestaña de configuración principal."""
        main = tk.Frame(parent, bg='#f4f7f9', padx=15, pady=10)
        main.pack(fill='both', expand=True)

        # Card 1: Server URL
        url_frame = tk.LabelFrame(
            main, text="📍 Portal BN", bg='#ffffff', fg='#C8102E',
            font=('Segoe UI', 10, 'bold'), padx=15, pady=10, relief='flat',
            highlightbackground='#e1e8ed', highlightthickness=1
        )
        url_frame.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 20))
        
        tk.Label(
            url_frame, text="URL del servidor para reporte automático:", 
            bg='#ffffff', font=('Segoe UI', 9), fg='#636e72'
        ).pack(anchor='w', pady=(0, 5))
        
        self.server_entry = tk.Entry(url_frame, font=('Segoe UI', 10), bg='#f8f9fa', relief='flat', highlightbackground='#e1e8ed', highlightthickness=1)
        self.server_entry.insert(0, "http://10.22.73.156:8080")
        self.server_entry.pack(fill='x', pady=5)

        # Card 2: File selection
        file_card = tk.LabelFrame(
            main, text="📂 Archivo Cronograma", bg='#ffffff', fg='#C8102E',
            font=('Segoe UI', 10, 'bold'), padx=15, pady=15, relief='flat',
            highlightbackground='#e1e8ed', highlightthickness=1
        )
        file_card.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(0, 20))
        
        self.file_label = tk.Label(
            file_card, text="Selecciona el archivo Excel de la campaña",
            bg='#ffffff', font=('Segoe UI', 10), fg='#636e72', anchor='w'
        )
        self.file_label.pack(side='left', expand=True, fill='x')
        
        tk.Button(
            file_card, text="📂 Cargar Excel", command=self._select_file,
            bg='#1a1c1e', fg='white', font=('Segoe UI', 9, 'bold'),
            relief='flat', padx=20, pady=8, cursor='hand2'
        ).pack(side='right')

        # Card 2b: Ruta de respaldo en red (NUEVA)
        backup_card = tk.LabelFrame(
            main, text="📁 Respaldo de Correos (Backup)", bg='#ffffff', fg='#C8102E',
            font=('Segoe UI', 10, 'bold'), padx=15, pady=12, relief='flat',
            highlightbackground='#e1e8ed', highlightthickness=1
        )
        backup_card.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(0, 20))

        # Descripción
        tk.Label(
            backup_card,
            text=("¿Dónde guardar los respaldos? Puedes apuntar a una carpeta de red/SharePoint"
                  " (recomendado para trabajo en equipo) o usar una carpeta local en esta PC."
                  " Si la ruta cambia, simplemente vuelve a seleccionarla."),
            bg='#ffffff', font=('Segoe UI', 8), fg='#636e72', wraplength=660, justify='left'
        ).pack(anchor='w', pady=(0, 8))

        # Fila 1: campo de texto + botón seleccionar red + botón local
        row_backup = tk.Frame(backup_card, bg='#ffffff')
        row_backup.pack(fill='x', pady=(0, 6))

        self.backup_path_var = tk.StringVar(value=self.backup_network_path)
        backup_entry = tk.Entry(
            row_backup, textvariable=self.backup_path_var,
            font=('Segoe UI', 9), bg='#f8f9fa', relief='flat',
            highlightbackground='#e1e8ed', highlightthickness=1
        )
        backup_entry.pack(side='left', expand=True, fill='x', padx=(0, 8))

        tk.Button(
            row_backup, text="📂 Seleccionar Carpeta de Red",
            command=self._select_backup_folder,
            bg='#1a1c1e', fg='white', font=('Segoe UI', 9, 'bold'),
            relief='flat', padx=12, pady=6, cursor='hand2'
        ).pack(side='left', padx=(0, 6))

        tk.Button(
            row_backup, text="🖥 Usar Local",
            command=self._use_local_backup_folder,
            bg='#2d3436', fg='white', font=('Segoe UI', 9),
            relief='flat', padx=10, pady=6, cursor='hand2'
        ).pack(side='left', padx=(0, 6))

        tk.Button(
            row_backup, text="✕ Limpiar",
            command=self._clear_backup_folder,
            bg='#636e72', fg='white', font=('Segoe UI', 9),
            relief='flat', padx=10, pady=6, cursor='hand2'
        ).pack(side='left')

        # Fila 2: botón de acción principal — generar respaldo
        row_gen = tk.Frame(backup_card, bg='#ffffff')
        row_gen.pack(fill='x', pady=(4, 0))

        tk.Label(
            row_gen,
            text="⚠ ¿Mitad de campaña o primera vez? Genera el respaldo de todos los correos ahora:",
            bg='#ffffff', font=('Segoe UI', 8, 'bold'), fg='#e67e22'
        ).pack(side='left', padx=(0, 10))

        tk.Button(
            row_gen, text="🔄 Generar Respaldo de Campaña Actual",
            command=self._generate_campaign_backup_all,
            bg='#C8102E', fg='white', font=('Segoe UI', 9, 'bold'),
            relief='flat', padx=14, pady=6, cursor='hand2'
        ).pack(side='left')

        # Indicador de estado
        self.backup_status_label = tk.Label(
            backup_card, text="", bg='#ffffff',
            font=('Segoe UI', 9, 'italic'), anchor='w'
        )
        self.backup_status_label.pack(fill='x', pady=(6, 0))
        self._update_backup_status()

        # Card 3: Validation (initially hidden)
        self.validation_frame = tk.LabelFrame(
            main, text="🔎 Auditoría de Columnas", bg='#ffffff', fg='#C8102E',
            font=('Segoe UI', 10, 'bold'), padx=15, pady=10, relief='flat',
            highlightbackground='#e1e8ed', highlightthickness=1
        )
        self.validation_frame.grid(row=4, column=0, columnspan=2, sticky='ew', pady=(0, 20))
        self.validation_frame.grid_remove()
        
        self.validation_text = scrolledtext.ScrolledText(
            self.validation_frame,
            height=6,
            font=('Consolas', 9),
            bg='#ffffff',
            wrap=tk.WORD
        )
        self.validation_text.pack(fill='both', expand=True, padx=10, pady=10)

        # Campaign info
        self.campaign_label = tk.Label(
            main, 
            text="", 
            bg='#f4f7f9',
            font=('Segoe UI', 14, 'bold'), 
            fg='#C8102E'
        )
        self.campaign_label.grid(row=5, column=0, columnspan=2, pady=(0, 15))

        # Mode selection (Card Style)
        mode_frame = tk.LabelFrame(
            main, 
            text="Modo de Generación", 
            bg='#ffffff',
            fg='#C8102E',
            font=('Segoe UI', 10, 'bold'),
            padx=15, pady=10,
            relief='flat',
            highlightbackground='#e1e8ed',
            highlightthickness=1
        )
        mode_frame.grid(row=6, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        
        self.mode_var = tk.StringVar(value='general')
        
        def on_mode_change():
            if self.mode_var.get() == 'reminder':
                self.cal_frame.pack(anchor='w', padx=15, pady=(0, 8))
            else:
                self.cal_frame.pack_forget()
            self._refresh_counts()

        tk.Radiobutton(
            mode_frame, 
            text="🔔 Inicio de Campaña (Todos los pagos programados)",
            variable=self.mode_var, 
            value='general', 
            bg='#ffffff',
            font=('Segoe UI', 10),
            activebackground='#ffffff',
            command=on_mode_change
        ).pack(anchor='w', padx=15, pady=8)
        
        tk.Radiobutton(
            mode_frame, 
            text="📌 Recordatorios (Elegir fecha objetivo)",
            variable=self.mode_var, 
            value='reminder', 
            bg='#ffffff',
            font=('Segoe UI', 10),
            activebackground='#ffffff',
            command=on_mode_change
        ).pack(anchor='w', padx=15, pady=8)

        # Calendar Selector
        self.cal_frame = tk.Frame(mode_frame, bg='#ffffff')
        tk.Label(self.cal_frame, text="Fecha Objetivo:", bg='#ffffff', font=('Segoe UI', 9)).pack(side='left', padx=(20, 5))
        
        def on_date_select(*args):
            self.selected_date = self.cal.get_date()
            self._refresh_counts()
            
        self.cal = DateEntry(self.cal_frame, width=12, background='darkblue',
                             foreground='white', borderwidth=2,
                             date_pattern='dd/mm/yyyy')
        self.cal.set_date(date.today() + timedelta(days=1))
        self.cal.pack(side='left')
        self.cal.bind("<<DateEntrySelected>>", on_date_select)
        
        # Initially hide calendar if general mode
        if self.mode_var.get() == 'reminder':
            self.cal_frame.pack(anchor='w', padx=15, pady=(0, 8))

        # Checkbox Modo Prueba
        self.test_mode_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            mode_frame,
            text="🧪 Modo Prueba (Previsualizar en Navegador sin usar Outlook)",
            variable=self.test_mode_var,
            bg='#ffffff', fg='#2980b9', font=('Segoe UI', 9, 'bold'),
            activebackground='#ffffff', activeforeground='#2980b9'
        ).pack(anchor='w', padx=15, pady=(10, 5))


        # Card 4: Automatización Completa
        self.auto_frame = tk.LabelFrame(
            main, 
            text="🤖 Automatización Completa (Segundo Plano)", 
            bg='#ffffff',
            fg='#C8102E',
            font=('Segoe UI', 10, 'bold'),
            padx=15, pady=15,
            relief='flat',
            highlightbackground='#e1e8ed',
            highlightthickness=1
        )
        self.auto_frame.grid(row=7, column=0, columnspan=2, sticky='ew', pady=(0, 20))

        self.auto_enabled_var = tk.BooleanVar(value=False)
        self.auto_hour_var = tk.StringVar(value="08:00")
        self.auto_send_var = tk.BooleanVar(value=True)

        auto_chk = tk.Checkbutton(
            self.auto_frame,
            text="Activar envío automático diario de recordatorios",
            variable=self.auto_enabled_var,
            bg='#ffffff', font=('Segoe UI', 10, 'bold'),
            command=self._on_auto_toggle
        )
        auto_chk.pack(anchor='w', pady=(0, 10))

        controls_f = tk.Frame(self.auto_frame, bg='#ffffff')
        controls_f.pack(fill='x', anchor='w', padx=25, pady=(0, 10))

        tk.Label(controls_f, text="Hora de ejecución (HH:MM):", bg='#ffffff', font=('Segoe UI', 9)).pack(side='left', padx=(0, 5))
        hour_ent = tk.Entry(controls_f, textvariable=self.auto_hour_var, width=8, font=('Segoe UI', 9), justify='center', highlightbackground='#e1e8ed', highlightthickness=1)
        hour_ent.pack(side='left', padx=(0, 20))
        hour_ent.bind("<FocusOut>", lambda e: self._save_config())

        send_chk = tk.Checkbutton(
            controls_f,
            text="Enviar directamente (si no se marca, se guardará en Borradores)",
            variable=self.auto_send_var,
            bg='#ffffff', font=('Segoe UI', 9),
            command=self._save_config
        )
        send_chk.pack(side='left')

        self.auto_status_label = tk.Label(
            self.auto_frame,
            text="Estado: Inactivo. Active la casilla para iniciar el servicio.",
            bg='#ffffff', fg='#636e72', font=('Segoe UI', 9, 'italic'),
            anchor='w'
        )
        self.auto_status_label.pack(fill='x', padx=25)

        # Progress
        progress_frame = tk.Frame(main, bg='#f4f7f9')
        progress_frame.grid(row=8, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            progress_frame, 
            variable=self.progress_var, 
            maximum=100
        )
        self.progress.pack(fill='x', pady=(0, 5))
        
        self.status_label = tk.Label(
            progress_frame, 
            text="Listo para comenzar", 
            bg='#f4f7f9',
            font=('Segoe UI', 9), 
            fg='#636e72',
            anchor='w'
        )
        self.status_label.pack(fill='x')

        # Generate button
        self.gen_btn = tk.Button(
            main, 
            text="⚡ GENERAR BORRADORES EN MI OUTLOOK",
            command=self._generate, 
            bg='#C8102E', 
            fg='white',
            font=('Segoe UI', 13, 'bold'), 
            relief='flat',
            padx=20, 
            pady=12, 
            cursor='hand2', 
            state='disabled'
        )
        self.gen_btn.grid(row=9, column=0, columnspan=2, sticky='ew')

        # Configurar pesos de las columnas
        main.columnconfigure(0, weight=1)

    def _build_preview_tab(self, parent):
        """Pestaña de preview de datos cargados."""
        frame = tk.Frame(parent, bg='#f4f7f9', padx=15, pady=15)
        frame.pack(fill='both', expand=True)

        tk.Label(
            frame, 
            text="Vista Previa del Cronograma Cargado", 
            bg='#f4f7f9',
            font=('Segoe UI', 12, 'bold'), 
            fg='#2d3436'
        ).pack(anchor='w', pady=(0, 10))

        self.preview_text = scrolledtext.ScrolledText(
            frame,
            font=('Consolas', 9),
            bg='#ffffff',
            wrap=tk.NONE
        )
        self.preview_text.pack(fill='both', expand=True)
        self.preview_text.insert('1.0', 'No hay datos cargados aún.\n\nSelecciona un archivo Excel en la pestaña de Configuración.')
        self.preview_text.config(state='disabled')

    def _build_log_tab(self, parent):
        """Pestaña de log de proceso."""
        frame = tk.Frame(parent, bg='#f4f7f9', padx=15, pady=15)
        frame.pack(fill='both', expand=True)

        tk.Label(
            frame, 
            text="Log del Proceso", 
            bg='#f4f7f9',
            font=('Segoe UI', 12, 'bold'), 
            fg='#2d3436'
        ).pack(anchor='w', pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(
            frame,
            font=('Consolas', 9),
            bg='#1e1e1e',
            fg='#d4d4d4',
            wrap=tk.WORD
        )
        self.log_text.pack(fill='both', expand=True)
        self._log("Sistema iniciado. Esperando carga de cronograma...")

    def _select_backup_folder(self):
        """Abre explorador de carpetas para seleccionar ruta de respaldo.
        Puede apuntar a: carpeta raiz (se creara BACKUP_CORREOS ahi) 
        o a la carpeta BACKUP_CORREOS ya existente directamente.
        """
        folder = filedialog.askdirectory(
            title="Seleccionar carpeta para backups (red, SharePoint o local)",
            mustexist=True
        )
        if folder:
            self.backup_network_path = folder
            self.backup_path_var.set(folder)
            self._update_backup_status()
            self._save_config()
            self._log(f"\U0001f4c1 Carpeta de respaldo configurada: {folder}")

    def _use_local_backup_folder(self):
        """Configura la carpeta local ./data como destino de respaldo."""
        import os as _os
        local_root = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'data')
        _os.makedirs(local_root, exist_ok=True)
        self.backup_network_path = local_root
        self.backup_path_var.set(local_root)
        self._update_backup_status()
        self._save_config()
        self._log(f"\U0001f5a5 Usando carpeta local como respaldo: {local_root}")

    def _clear_backup_folder(self):
        """Limpia la ruta de respaldo — vuelve al fallback (junto al cronograma)."""
        self.backup_network_path = ''
        self.backup_path_var.set('')
        self._update_backup_status()
        self._save_config()
        self._log("\U0001f4c1 Ruta de respaldo limpiada. Se usara la carpeta del cronograma como referencia.")

    def _update_backup_status(self):
        """Actualiza el indicador de estado de conexion a la carpeta de respaldo."""
        if not hasattr(self, 'backup_status_label'):
            return
        path = self.backup_network_path
        if not path:
            self.backup_status_label.config(
                text="\u26a0 No configurada \u2014 el respaldo se guardara junto al Excel del cronograma.",
                fg='#e67e22'
            )
        elif os.path.isdir(path):
            backup_root = os.path.join(path, 'BACKUP_CORREOS')
            if os.path.isdir(backup_root):
                self.backup_status_label.config(
                    text="\u2705 Conectada \u2014 carpeta BACKUP_CORREOS ya existe en la ruta configurada.",
                    fg='#27ae60'
                )
            else:
                self.backup_status_label.config(
                    text="\u2705 Ruta accesible \u2014 BACKUP_CORREOS se creara al generar el primer correo.",
                    fg='#27ae60'
                )
        else:
            self.backup_status_label.config(
                text="\u274c Ruta no accesible \u2014 verifica la conexion a la red o SharePoint.",
                fg='#e74c3c'
            )

    def _generate_campaign_backup_all(self):
        """Genera (o regenera) el respaldo de TODAS las agencias sin enviar nada.
        
        Util para:
        - Mitad de campana: bootstrapping del sistema de backup
        - Cambio de ruta: regenerar en nueva ubicacion
        - Usuario externo sin acceso al enviado original
        """
        if self.df_cronograma is None or self.df_cronograma.empty:
            messagebox.showwarning(
                "Sin cronograma",
                "Primero debes cargar el Excel de la campana desde la pestana Configuracion."
            )
            return

        backup_dir = self._get_backup_dir()
        self._log(f"\n\U0001f504 Iniciando generacion de respaldo para TODAS las agencias...")
        self._log(f"   Destino: {backup_dir}")
        self._log(f"   Campana: {self.campaign}\n")

        agencies = sorted(self.df_cronograma['Agencia'].dropna().unique().tolist())
        ok, fail = 0, 0

        for agency in agencies:
            try:
                df_ag = self.df_cronograma[self.df_cronograma['Agencia'] == agency].copy()
                if df_ag.empty:
                    continue

                recipients = self._get_recipients(agency)
                has_juntos = ('Juntos_Usuarios' in df_ag.columns and
                              df_ag['Juntos_Usuarios'].notna().any() and
                              (df_ag['Juntos_Usuarios'].fillna(0) > 0).any())
                has_contigo = ('Contigo_Usuarios' in df_ag.columns and
                               df_ag['Contigo_Usuarios'].notna().any() and
                               (df_ag['Contigo_Usuarios'].fillna(0) > 0).any())

                table = self._build_html_table(df_ag, has_juntos, has_contigo, is_reminder=False)
                html = (f'<div style="font-family:Calibri;font-size:11pt;color:black;">'
                        f'<p>Estimado(a) buenos dias,<br><b>BN - {agency}</b></p>'
                        f'<p>Segun el cronograma de pagos para la campana <b>{self.campaign}</b>,'
                        f' se adjunta el detalle correspondiente a su agencia.</p>'
                        f'<br>{table}<br></div>')
                subject = f'Programacion de Fondos - BN {agency} - {self.campaign}'

                # 1. Guardar HTML + metadata.json
                self._save_complete_backup(agency, subject, recipients, html)

                # 2. Guardar Excel filtrado en subcarpeta de agencia
                try:
                    import openpyxl
                    from openpyxl import Workbook
                    wb_src = openpyxl.load_workbook(self.cronograma_path)
                    ws_src = wb_src.active
                    header_row = None
                    for row in ws_src.iter_rows():
                        for cell in row:
                            if str(cell.value or '').strip().upper() in ['AGENCIA', 'AG', 'AGENCIAS']:
                                header_row = cell.row
                                break
                        if header_row:
                            break
                    if header_row:
                        wb_out = Workbook()
                        ws_out = wb_out.active
                        headers = [cell.value for cell in ws_src[header_row]]
                        ws_out.append(headers)
                        ag_col_idx = next(
                            (i for i, h in enumerate(headers)
                             if str(h or '').strip().upper() in ['AGENCIA', 'AG', 'AGENCIAS']),
                            None
                        )
                        if ag_col_idx is not None:
                            for row in ws_src.iter_rows(min_row=header_row + 1, values_only=True):
                                if row[ag_col_idx] and str(row[ag_col_idx]).strip().upper() == agency.upper():
                                    ws_out.append(list(row))
                        agency_dir = self._get_agency_backup_dir(agency)
                        orig_name = os.path.basename(self.cronograma_path)
                        wb_out.save(os.path.join(agency_dir, orig_name))
                except Exception as xe:
                    logging.debug(f"Excel para {agency}: {xe}")

                self._log(f"   \u2713 {agency}")
                ok += 1

            except Exception as e:
                self._log(f"   \u26a0 {agency}: {e}")
                fail += 1

        self._log(f"\n\u2705 Respaldo completado: {ok} agencias OK, {fail} con error.")
        messagebox.showinfo(
            "Respaldo generado",
            f"Respaldo generado para {ok} agencia(s).\n\nUbicacion:\n{backup_dir}\n\n"
            f"Cualquier persona con acceso a esa carpeta puede ahora "
            f"generar recordatorios en hilo."
        )

    def _select_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Cronograma ETV",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if path:
            self._load_cronograma_from_path(path)

    def _load_cronograma_from_path(self, path):
        """Carga el cronograma desde una ruta específica y actualiza la UI."""
        self._log(f"\n{'='*60}")
        self._log(f"Cargando: {os.path.basename(path)}")
        
        try:
            # Cargar con sistema flexible
            df, missing = ExcelLoader.load_flexible(path)
            
            self.df_cronograma = df
            self.cronograma_path = path
            self.missing_columns = missing
            
            # Detectar campaña
            self._detect_campaign(path)
            
            # Actualizar UI
            self.file_label.config(text=os.path.basename(path), fg='#27ae60')
            self.campaign_label.config(text=f"📋 {self.campaign}")
            
            # Mostrar validación
            self._show_validation_status(df, missing)
            
            # Actualizar preview
            self._update_preview(df)
            
            # Guardar persistencia
            self._save_config()
            
            # Habilitar botón si todo OK
            agencies = self._get_agencies()
            if len(agencies) > 0:
                self.gen_btn.config(state='normal')
                self.status_label.config(
                    text=f"✓ Cronograma válido: {len(agencies)} agencias detectadas",
                    fg='#27ae60'
                )
                self._refresh_counts()
                self._log(f"✓ Carga exitosa: {len(df)} filas, {len(agencies)} agencias")
            else:
                self.gen_btn.config(state='disabled')
                self.status_label.config(
                    text="⚠ No se detectaron agencias válidas",
                    fg='#e74c3c'
                )
                self._log("⚠ Advertencia: No se encontraron agencias en el archivo")
                
        except Exception as e:
            self._log(f"✗ ERROR: {str(e)}")
            if hasattr(self, 'root'):
                messagebox.showerror("Error al Cargar", f"No se pudo cargar el cronograma:\n\n{e}")

    def _detect_campaign(self, path):
        """Detecta campaña desde el nombre del archivo."""
        basename = os.path.basename(path).upper()
        code = None
        
        for c, full_name in CAMPAIGN_CODES.items():
            if re.search(rf'\b{c}\b', basename, re.IGNORECASE):
                code = c
                break
        
        year_match = re.search(r'\b(20\d{2})\b', basename)
        year = year_match.group(1) if year_match else str(datetime.now().year)
        
        self.campaign_code = code or '??'
        self.campaign = f"{CAMPAIGN_CODES.get(self.campaign_code, 'CAMPAÑA')} {year}"
        
        if not code:
            self._log(f"⚠ No se pudo detectar código de campaña en '{basename}'")
            self._log(f"  Se usará: {self.campaign}")

    def _show_validation_status(self, df, missing):
        """Muestra estado de validación en la UI."""
        self.validation_frame.grid()
        self.validation_text.delete('1.0', tk.END)
        
        # Resumen
        status = "✓ VALIDACIÓN EXITOSA\n" if not missing else "⚠ VALIDACIÓN CON ADVERTENCIAS\n"
        self.validation_text.insert(tk.END, status, 'header')
        self.validation_text.insert(tk.END, "="*50 + "\n\n")
        
        # Columnas encontradas
        required_found = [col for col in REQUIRED_COLUMNS.keys() if col in df.columns]
        self.validation_text.insert(tk.END, f"Columnas críticas encontradas ({len(required_found)}):\n", 'bold')
        for col in required_found:
            self.validation_text.insert(tk.END, f"  ✓ {col}\n", 'success')
        
        # Columnas faltantes
        if missing:
            self.validation_text.insert(tk.END, f"\nColumnas faltantes ({len(missing)}):\n", 'bold')
            for col in missing:
                self.validation_text.insert(tk.END, f"  ✗ {col}\n", 'warning')
        
        # Estadísticas
        self.validation_text.insert(tk.END, f"\nEstadísticas:\n", 'bold')
        self.validation_text.insert(tk.END, f"  • Filas totales: {len(df)}\n")
        self.validation_text.insert(tk.END, f"  • Agencias únicas: {df['Agencia'].nunique()}\n")
        
        # Configurar tags de colores
        self.validation_text.tag_config('header', font=('Consolas', 9, 'bold'))
        self.validation_text.tag_config('bold', font=('Consolas', 9, 'bold'))
        self.validation_text.tag_config('success', foreground='#27ae60')
        self.validation_text.tag_config('warning', foreground='#e67e22')

    def _update_preview(self, df):
        """Actualiza preview de datos."""
        self.preview_text.config(state='normal')
        self.preview_text.delete('1.0', tk.END)
        
        # Mostrar info general
        self.preview_text.insert(tk.END, f"CRONOGRAMA CARGADO\n")
        self.preview_text.insert(tk.END, f"="*80 + "\n\n")
        self.preview_text.insert(tk.END, f"Filas: {len(df)} | Columnas: {len(df.columns)}\n\n")
        
        # Listar agencias
        agencies = sorted(df['Agencia'].unique())
        self.preview_text.insert(tk.END, f"AGENCIAS DETECTADAS ({len(agencies)}):\n")
        self.preview_text.insert(tk.END, "-"*80 + "\n")
        for i, ag in enumerate(agencies, 1):
            count = len(df[df['Agencia'] == ag])
            self.preview_text.insert(tk.END, f"{i:2d}. {ag:40s} ({count:3d} pagos)\n")
        
        # Muestra de datos
        self.preview_text.insert(tk.END, f"\n\nMUESTRA DE DATOS (primeras 5 filas):\n")
        self.preview_text.insert(tk.END, "="*80 + "\n")
        self.preview_text.insert(tk.END, df.head().to_string() + "\n")
        
        self.preview_text.config(state='disabled')

    def _get_agencies(self):
        """Devuelve lista de agencias únicas del cronograma."""
        if self.df_cronograma is None:
            return []
        return sorted(self.df_cronograma['Agencia'].unique().tolist())

    def _update_status(self, text):
        """Actualiza el label de estado en la UI."""
        self.root.after(0, lambda: self.status_label.config(text=text))

    def _update_progress(self, val, status):
        """Actualiza progreso y estado."""
        self.root.after(0, lambda: self.progress_var.set(val))
        self.root.after(0, lambda: self.status_label.config(text=status))

    def _refresh_counts(self):
        """Actualiza el contador de agencias según el modo seleccionado."""
        if self.df_cronograma is None: return
        
        mode = self.mode_var.get()
        if mode == 'general':
            agencies = self._get_agencies()
            self.status_label.config(
                text=f"✓ Modo General: {len(agencies)} agencias detectadas",
                fg='#27ae60'
            )
        else:
            tomorrow = self.selected_date
            tomorrow_str = tomorrow.strftime('%d/%m/%Y')
            
            # Usar la fecha calculada (mínima entre pago y entrega)
            df = self.df_cronograma.copy()
            
            if 'Target_Date_Calc' in df.columns:
                df_tomorrow = df[df['Target_Date_Calc'] == tomorrow]
            else:
                # Fallback por si la columna no existe
                date_col = 'Fecha_Entrega' if 'Fecha_Entrega' in df.columns else 'Fecha_Pago'
                df['Target_Date'] = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
                df_tomorrow = df[df['Target_Date'].dt.date == tomorrow]
            
            agencies = sorted(df_tomorrow['Agencia'].unique().tolist()) if not df_tomorrow.empty else []
            
            if len(agencies) > 0:
                self.status_label.config(
                    text=f"✓ Recordatorios (Fecha Prox): {len(agencies)} agencias para {tomorrow_str}",
                    fg='#27ae60'
                )
            else:
                self.status_label.config(
                    text=f"⚠ Sin eventos programados para {tomorrow_str}",
                    fg='#e67e22'
                )

    def _generate(self):
        # No cambiar de pestaña automaticamente
        self.gen_btn.config(state='disabled')
        
        mode = self.mode_var.get()
        if mode == 'reminder':
            # Pre-filtrar agencias por fecha para los recordatorios
            target_date = self.selected_date
            df = self.df_cronograma.copy()
            if 'Target_Date_Calc' in df.columns:
                df_filt = df[df['Target_Date_Calc'] == target_date]
            else:
                date_col = 'Fecha_Entrega' if 'Fecha_Entrega' in df.columns else 'Fecha_Pago'
                df['Tmp_Date'] = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
                df_filt = df[df['Tmp_Date'].dt.date == target_date]
            
            agencies_to_show = sorted(df_filt['Agencia'].unique().tolist()) if not df_filt.empty else []
            if not agencies_to_show:
                messagebox.showinfo("Sin pagos", f"No se encontraron pagos programados para el {target_date.strftime('%d/%m/%Y')}")
                self.gen_btn.config(state='normal')
                return
        else:
            # Para correo general, mostrar todas las agencias detectadas
            agencies_to_show = self._get_agencies()

        if not agencies_to_show:
            messagebox.showwarning("Sin agencias", "No hay agencias disponibles para procesar.")
            self.gen_btn.config(state='normal')
            return

        # Pedir selección de agencias
        selected = self._ask_agencies_selection(agencies_to_show)
        if not selected:
            self._log("Proceso cancelado por el usuario.")
            self.gen_btn.config(state='normal')
            return

        self.error_log = []
        self._log(f"\n{'='*60}")
        self._log(f"INICIANDO GENERACIÓN - {len(selected)} agencias")
        self._log(f"Modo: {'Recordatorios' if mode == 'reminder' else 'Correos Principales'}")
        
        # Ejecutar en hilo para no congelar la UI
        threading.Thread(target=self._generate_thread, args=(selected,), daemon=True).start()

    def _ask_agencies_selection(self, agencies):
        """Muestra ventana modal para seleccionar agencias a procesar."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Seleccionar Agencias")
        dialog.geometry("400x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar relativo a la principal
        dx = self.root.winfo_x() + 100
        dy = self.root.winfo_y() + 50
        dialog.geometry(f"+{dx}+{dy}")

        main_frame = tk.Frame(dialog, bg='#f4f7f9', padx=10, pady=10)
        main_frame.pack(fill='both', expand=True)

        tk.Label(main_frame, text="Marque las agencias a generar:", bg='#f4f7f9', font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(0,10))

        # Canvas con scroll para la lista
        list_frame = tk.Frame(main_frame, bg='white', relief='sunken', bd=1)
        list_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(list_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='white')

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # MouseWheel binding for the popup
        def _on_mousewheel_popup(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except Exception:
                pass
        self.root.bind_all("<MouseWheel>", _on_mousewheel_popup)

        vars_dict = {}
        # Ordenar alfabéticamente
        sorted_ag = sorted(agencies)
        for ag in sorted_ag:
            v = tk.BooleanVar(value=True)
            cb = tk.Checkbutton(scroll_frame, text=ag, variable=v, bg='white', anchor='w')
            cb.pack(fill='x', padx=5, pady=2)
            vars_dict[ag] = v

        btn_frame = tk.Frame(main_frame, bg='#f4f7f9', pady=10)
        btn_frame.pack(fill='x')

        result = []
        
        def _cleanup():
            # Restaurar el hook del ratón de la ventana principal
            self.root.bind_all("<MouseWheel>", self._on_main_mousewheel)
            dialog.destroy()
            
        def on_confirm():
            for ag, v in vars_dict.items():
                if v.get():
                    result.append(ag)
            _cleanup()

        def on_toggle_all():
            val = not any(v.get() for v in vars_dict.values())
            for v in vars_dict.values():
                v.set(val)

        tk.Button(btn_frame, text="Marcar/Desmarcar Todo", command=on_toggle_all, bg='#636e72', fg='white', relief='flat', padx=10).pack(side='left')
        tk.Button(btn_frame, text="Confirmar Selección", command=on_confirm, bg='#C8102E', fg='white', relief='flat', padx=20, font=('Segoe UI', 9, 'bold')).pack(side='right')

        dialog.protocol("WM_DELETE_WINDOW", _cleanup)

        self.root.wait_window(dialog)
        return result

    def _generate_thread(self, agencies):
        test_mode = getattr(self, 'test_mode_var', tk.BooleanVar(value=False)).get()
        
        # COM requiere inicialización explícita en cada hilo secundario
        com_initialized = False
        if not test_mode:
            try:
                import pythoncom
                pythoncom.CoInitialize()
                com_initialized = True
            except Exception:
                pass
            try:
                import win32com.client
                outlook = win32com.client.Dispatch("Outlook.Application")
                self._log("✓ Conexión con Outlook establecida")
            except Exception as e:
                self._log(f"✗ ERROR: No se pudo conectar a Outlook: {e}")
                self._update_status(f"ERROR: No se pudo conectar a Outlook")
                self.root.after(0, lambda: messagebox.showerror(
                    "Error de Conexión", 
                    "No se pudo conectar a Microsoft Outlook.\n\nAsegúrate de tener Outlook instalado y configurado en este equipo, o activa la casilla de 'Modo Prueba' para revisar los correos en tu navegador."
                ))
                self.gen_btn.config(state='normal')
                if com_initialized:
                    pythoncom.CoUninitialize()
                return
        else:
            outlook = None
            self._log("🧪 MODO PRUEBA ACTIVO: Generación local de HTML")
            if hasattr(self, '_test_html_opened'):
                delattr(self, '_test_html_opened')

        mode = self.mode_var.get()
        # Nota: 'agencies' ya viene filtrado desde _generate()
        
        self._log(f"\nProcesando {len(agencies)} agencias...")
        self._log("-"*60)

        # preparar contenedor de resultados para reporte y resumen
        results = []

        for i, agency in enumerate(agencies):
            pct = ((i + 1) / len(agencies)) * 100
            self._update_progress(pct, f"[{i+1}/{len(agencies)}] {agency}...")
            self._log(f"\n{i+1}. {agency}")

            try:
                if mode == 'general':
                    result = self._create_general_draft(outlook, agency)
                else:
                    result = self._create_reminder_draft(outlook, agency)
                
                results.append(result)

                # log differently depending on status
                if result.get('status') == 'GENERADO':
                    self._log(f"   ✓ {result['status']} → {result.get('to','')}")
                else:
                    err = result.get('error','')
                    self._log(f"   ✗ {result.get('status')} {('('+err+')' if err else '')}")
                    self.error_log.append({'agency': agency, 'error': err})
                
            except Exception as e:
                error_msg = str(e)
                results.append({'agency': agency, 'status': 'ERROR', 'to': '', 'error': error_msg})
                self._log(f"   ✗ ERROR: {error_msg}")
                self.error_log.append({'agency': agency, 'error': error_msg})

        # Resumen
        ok = sum(1 for r in results if r['status'] == 'GENERADO')
        err = sum(1 for r in results if r['status'] == 'ERROR')
        
        self._log(f"\n{'='*60}")
        self._log(f"PROCESO COMPLETADO")
        self._log(f"✓ Generados: {ok}")
        self._log(f"✗ Errores: {err}")
        
        # Mostrar errores específicos
        if self.error_log:
            self._log(f"\nDETALLE DE ERRORES:")
            for item in self.error_log:
                self._log(f"  • {item['agency']}: {item['error']}")
        
        self._update_status(f"✅ Completado: {ok} generados, {err} errores")

        # Reportar al portal
        self._report_to_portal(results)

        # Mensaje final
        msg = f"Se generaron {ok} borradores en tu Outlook.\n"
        if err > 0:
            msg += f"\n⚠ {err} agencias tuvieron errores.\n"
            msg += "Revisa la pestaña 'Log' para más detalles.\n"
        msg += "\nRevisa tu carpeta de Borradores para enviar los correos."
        
        self.root.after(0, lambda: messagebox.showinfo("Proceso Completado", msg))
        self.root.after(0, lambda: self.gen_btn.config(state='normal'))
        
        # Liberar COM al terminar el hilo
        if com_initialized:
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _create_general_draft(self, outlook, agency):
        """Crea borrador de correo general."""
        df_ag = self.df_cronograma[self.df_cronograma['Agencia'] == agency]
        recipients = self._get_recipients(agency)
        
        # Detectar programas sociales activos
        has_juntos = False
        has_contigo = False
        
        if 'Juntos_Usuarios' in df_ag.columns:
            has_juntos = df_ag['Juntos_Usuarios'].notna().any() and (df_ag['Juntos_Usuarios'].fillna(0) > 0).any()
        
        if 'Contigo_Usuarios' in df_ag.columns:
            has_contigo = df_ag['Contigo_Usuarios'].notna().any() and (df_ag['Contigo_Usuarios'].fillna(0) > 0).any()

        # Fecha de inicio desde Fecha_Entrega (no Fecha_Pago)
        meses_es = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        fecha_inicio = ''
        try:
            date_col_for_inicio = 'Fecha_Entrega' if 'Fecha_Entrega' in df_ag.columns else 'Fecha_Pago'
            fechas = pd.to_datetime(df_ag[date_col_for_inicio], errors='coerce').dropna()
            if not fechas.empty:
                fd = fechas.min()
                fecha_inicio = f"{fd.day:02d} DE {meses_es[fd.month-1]} DE {fd.year}"
        except:
            fecha_inicio = 'PRÓXIMAMENTE'

        html = self._build_html_body(agency, fecha_inicio, df_ag, has_juntos, has_contigo)
        subject = f'Programación de Fondos – BN {agency} – {self.campaign}'

        if outlook is None:
            return self._save_test_html(agency, subject, html)

        mail = outlook.CreateItem(0)
        mail.Subject = subject
        
        # Cargar firma silenciosa si está disponible
        signature = self._get_signature_for_mail(mail)
        if signature:
            if "<body>" in signature.lower():
                parts = re.split(r'(<body>)', signature, flags=re.IGNORECASE)
                mail.HTMLBody = parts[0] + parts[1] + html + "<br><br>" + parts[2]
            else:
                mail.HTMLBody = html + "<br><br>" + signature
        else:
            mail.HTMLBody = html
            
        mail.To = recipients['to']
        if recipients['cc']:
            mail.CC = recipients['cc']
        
        # Adjuntar Excel filtrado usando el helper
        self._attach_filtered_excel_to_mail(mail, agency)
        
        # Guardar respaldo completo (HTML + metadata JSON) para recuperación multi-usuario
        self._save_complete_backup(agency, subject, recipients, html)
        mail.Save()

        return {'agency': agency, 'status': 'GENERADO', 'to': recipients['to']}

    def _create_reminder_draft(self, outlook, agency, auto_send=False):
        """Crea borrador o envía recordatorio de correo respondiendo al hilo original o usando backup."""
        target_date = self.selected_date
        df_ag = self.df_cronograma[self.df_cronograma['Agencia'] == agency]

        if 'Target_Date_Calc' in df_ag.columns:
            df_tomorrow = df_ag[df_ag['Target_Date_Calc'] == target_date]
        else:
            date_col = 'Fecha_Entrega' if 'Fecha_Entrega' in df_ag.columns else 'Fecha_Pago'
            df_ag = df_ag.copy()
            df_ag['Target_Date'] = pd.to_datetime(df_ag[date_col], errors='coerce', dayfirst=True)
            df_tomorrow = df_ag[df_ag['Target_Date'].dt.date == target_date]

        if df_tomorrow.empty:
            return {'agency': agency, 'status': 'ERROR', 'to': '', 'error': 'Sin pagos en fecha seleccionada'}

        recipients = self._get_recipients(agency)
        has_juntos = ('Juntos_Usuarios' in df_tomorrow.columns and 
                      df_tomorrow['Juntos_Usuarios'].notna().any() and 
                      (df_tomorrow['Juntos_Usuarios'].fillna(0) > 0).any())
        has_contigo = ('Contigo_Usuarios' in df_tomorrow.columns and 
                       df_tomorrow['Contigo_Usuarios'].notna().any() and 
                       (df_tomorrow['Contigo_Usuarios'].fillna(0) > 0).any())
        
        # Lógica de Fecha (Viernes -> Lunes)
        dias_dif = (target_date - date.today()).days
        meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        dias = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']
        dia_semana = dias[target_date.weekday()]
        mes = meses[target_date.month - 1]
        fecha_fmt = f"{target_date.day:02d} de {mes} de {target_date.year}"

        if dias_dif <= 1:
            fecha_texto = "el día de mañana"
        else:
            fecha_texto = f"el {dia_semana}"

        table = self._build_html_table(df_tomorrow, has_juntos, has_contigo, is_reminder=True)
        
        cuerpo_html = f'''<div style="font-family:Calibri;font-size:11pt;color:black;">
<p>Estimado(a) buenos días,<br><b>BN – {agency}</b></p>
<p>Se le recuerda que para {fecha_texto} <b><span style="background:yellow;mso-highlight:yellow">{fecha_fmt}</span></b>, se debe programar la entrega de fondos a la ETV indicada.</p>
<p><b><span style="font-size:12.0pt;color:#C00000">Se adjunta archivo en formato Excel con la programación detallada de dicha campaña.</span></b></p>
<p><b><i><span style="font-size:12.0pt;background:aqua;mso-highlight:aqua">Obs: En el caso de la ETV PROSEGUR, se deberá imprimir y entregar las planillas.</span></i></b></p>
<p>De antemano se agradece por el cumplimiento del calendario asignado para esta campaña ({self.campaign}).</p>
<br>{table}<br></div>'''

        new_subject = f'RECORDATORIO: Entrega de Fondos – BN {agency} – {self.campaign}'

        if outlook is None:
            # En modo prueba, intentar leer el backup y previsualizar
            backup_html = ""
            try:
                backup_path = self._find_backup_file(agency)
                if backup_path and os.path.exists(backup_path):
                    with open(backup_path, 'r', encoding='utf-8') as f:
                        backup_html = f.read()
            except:
                pass
            
            full_body = cuerpo_html
            if backup_html:
                full_body += f"<br><hr><div style='font-family:Calibri;font-size:10pt;color:#555;'><b>De:</b> Caja y Valores BN<br><b>Asunto:</b> Programación de Fondos – BN {agency} – {self.campaign}</div><br>{backup_html}"
            return self._save_test_html(agency, new_subject, full_body)

        # ---------------------------------------------------------
        # BÚSQUEDA DEL HILO ORIGINAL (Enviados -> Entrada)
        # ---------------------------------------------------------
        recipients_to = recipients.get('to', '') if isinstance(recipients, dict) else ''
        original_mail = self._find_original_mail(outlook, agency, recipients_to)
        # ---------------------------------------------------------

        mail = None
        if original_mail:
            try:
                mail = original_mail.ReplyAll()
                # Reemplazar RV: / Re: / FW: por RECORDATORIO:
                orig_subj = mail.Subject if mail.Subject else original_mail.Subject
                clean_subj = re.sub(r'^(RV:|RE:|FW:|R:|F:|RES:|Re:|Fw:)\s*', '', orig_subj, flags=re.IGNORECASE).strip()
                mail.Subject = f'RECORDATORIO: {clean_subj}'
                
                # Incorporar firma si está disponible
                signature = self._get_signature_for_mail(mail)
                if signature:
                    if "<body>" in signature.lower():
                        parts = re.split(r'(<body>)', signature, flags=re.IGNORECASE)
                        mail.HTMLBody = parts[0] + parts[1] + cuerpo_html + "<br><br>" + parts[2]
                    else:
                        mail.HTMLBody = cuerpo_html + "<br><br>" + signature
                else:
                    mail.HTMLBody = cuerpo_html + mail.HTMLBody
                
                # Copiar SOLO adjuntos .xlsx del correo padre al borrador de respuesta
                import tempfile, shutil
                tmp_dir = tempfile.mkdtemp(prefix='etv_reply_')
                try:
                    for att in original_mail.Attachments:
                        try:
                            fname = att.FileName
                            if not fname.lower().endswith('.xlsx'):
                                continue  # Saltar firmas, imagenes, etc.
                            tmp_path = os.path.join(tmp_dir, fname)
                            att.SaveAsFile(tmp_path)
                            mail.Attachments.Add(tmp_path)
                            self._log(f"   (📎 Adjunto copiado del original: {fname})")
                        except Exception as att_err:
                            self._log(f"   (⚠ No se pudo copiar adjunto: {att_err})")
                except Exception:
                    pass
                
                # Limpiar destinatarios del ReplyAll — quitar jefe de Caja y externos
                try:
                    EXCLUDED_EMAILS = {'hbazan@bn.com.pe'}
                    INTERNAL_DOMAIN = 'bn.com.pe'

                    def _get_smtp(recip):
                        """Obtiene la direccion SMTP real aunque sea un contacto Exchange."""
                        try:
                            ae = recip.AddressEntry
                            if ae and ae.AddressEntryUserType in (0, 10):  # 0=Exchange User, 10=SMTP
                                eu = ae.GetExchangeUser()
                                if eu and eu.PrimarySmtpAddress:
                                    return eu.PrimarySmtpAddress.lower()
                        except Exception:
                            pass
                        try:
                            return recip.Address.lower()
                        except Exception:
                            return ''

                    to_remove = []
                    for i in range(1, mail.Recipients.Count + 1):
                        try:
                            addr = _get_smtp(mail.Recipients.Item(i))
                            self._log(f"   [Recip] {i}: {addr}")
                            if addr in EXCLUDED_EMAILS:
                                to_remove.append(i)
                                self._log(f"   [Recip] → Eliminado (jefe): {addr}")
                            elif INTERNAL_DOMAIN not in addr and '@' in addr:
                                to_remove.append(i)
                                self._log(f"   [Recip] → Eliminado (externo): {addr}")
                        except Exception:
                            pass

                    for i in reversed(to_remove):
                        mail.Recipients.Remove(i)
                    mail.Recipients.ResolveAll()
                except Exception as recip_err:
                    self._log(f"   (⚠ No se pudo limpiar destinatarios: {recip_err})")
                
                self._log(f"   (✓ Respondiendo al hilo original: {original_mail.Subject})")
            except Exception as e_reply:
                self._log(f"   Advertencia: Falló ReplyAll ({e_reply}), creando nuevo.")
                mail = None

        if not mail:
            self._log("   (⚠ Correo original no encontrado en Outlook — buscando respaldo completo en red)")
            
            # Intentar cargar el backup completo (metadata JSON + HTML + Excel)
            backup_data = self._load_backup_metadata(agency)
            backup_html = ""
            backup_meta = None
            backup_excel = None
            sent_by = 'Caja y Valores BN'
            saved_at_fmt = '(fecha desconocida)'
            original_to = recipients.get('to', '')
            original_subject = f'Programación de Fondos – BN {agency} – {self.campaign}'
            
            if backup_data:
                backup_html = backup_data.get('html', '')
                backup_meta = backup_data.get('metadata')
                backup_excel = backup_data.get('excel_path')
                if backup_meta:
                    sent_by = backup_meta.get('sent_by', 'Caja y Valores BN')
                    raw_at = backup_meta.get('saved_at', '')
                    try:
                        dt = datetime.fromisoformat(raw_at)
                        saved_at_fmt = dt.strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        saved_at_fmt = raw_at or '(fecha desconocida)'
                    original_to = backup_meta.get('to', original_to)
                    original_subject = backup_meta.get('subject', original_subject)
                    self._log(f"   (✓ Respaldo completo encontrado — enviado por '{sent_by}' el {saved_at_fmt})")
                else:
                    self._log("   (✓ Respaldo HTML encontrado — sin metadata)")
            else:
                # Fallback: método legacy (busca principal_AGENCIA.html plano)
                try:
                    legacy_path = self._find_backup_file(agency)
                    if legacy_path and os.path.exists(legacy_path):
                        with open(legacy_path, 'r', encoding='utf-8') as f:
                            backup_html = f.read()
                        self._log(f"   (✓ Respaldo legacy encontrado: {os.path.basename(legacy_path)})")
                    else:
                        self._log(f"   (⚠ Sin respaldo para {agency} — se creará recordatorio independiente)")
                except Exception as re_err:
                    self._log(f"   (⚠ Error leyendo respaldo: {re_err})")
                
            mail = outlook.CreateItem(0)
            mail.Subject = new_subject
            mail.To = original_to or recipients.get('to', '')
            if recipients.get('cc'):
                mail.CC = recipients['cc']
            
            # Construir cuerpo simulando cadena de respuesta al correo original
            signature = self._get_signature_for_mail(mail)
            simulated_reply = cuerpo_html
            if backup_html:
                simulated_reply += f'''<br><br>
<hr align="center" size="2" width="100%" tabindex="-1">
<div style="font-family:Calibri,sans-serif;font-size:10pt;color:#555;direction:ltr;">
<b>De:</b> {sent_by}<br>
<b>Enviado:</b> {saved_at_fmt}<br>
<b>Para:</b> {original_to}<br>
<b>Asunto:</b> {original_subject}<br>
</div>
<br>
{backup_html}'''
            
            if signature:
                if "<body>" in signature.lower():
                    parts = re.split(r'(<body>)', signature, flags=re.IGNORECASE)
                    mail.HTMLBody = parts[0] + parts[1] + simulated_reply + "<br><br>" + parts[2]
                else:
                    mail.HTMLBody = simulated_reply + "<br><br>" + signature
            else:
                mail.HTMLBody = simulated_reply
                
            # Adjuntar Excel: usar el del backup si existe, si no regenerar
            if backup_excel and os.path.exists(backup_excel):
                try:
                    mail.Attachments.Add(os.path.abspath(backup_excel))
                    self._log(f"   (📎 Excel adjunto desde respaldo: {os.path.basename(backup_excel)})")
                except Exception as att_err:
                    self._log(f"   (⚠ No se pudo adjuntar Excel del backup, regenerando: {att_err})")
                    self._attach_filtered_excel_to_mail(mail, agency)
            else:
                self._attach_filtered_excel_to_mail(mail, agency)

        if auto_send:
            mail.Send()
            self._log(f"   ✓ [AUTOMÁTICO] Correo ENVIADO directamente a {mail.To or recipients['to']}")
            return {'agency': agency, 'status': 'ENVIADO', 'to': mail.To or recipients['to']}
        else:
            mail.Save()
            return {'agency': agency, 'status': 'GENERADO', 'to': mail.To or recipients['to']}

    def _split_search_emails(self, recipients_to):
        import re
        try:
            return [e.strip().upper() for e in re.split(r"[;,]", str(recipients_to or "")) if e.strip()]
        except Exception:
            return []

    def _mail_matches_original_thread(self, mail, agency, recipients_to):
        try:
            if mail.Class != 43:  # olMail
                return False

            subj = (mail.Subject or "").upper().strip()
            if self.campaign.upper() not in subj:
                return False

            if subj.startswith("RECORDATORIO"):
                return False

            agency_up = str(agency).strip().upper()
            search_emails = self._split_search_emails(recipients_to)

            match_agency = agency_up in subj

            mail_to = ""
            try:
                mail_to = (mail.To or "").upper()
            except Exception:
                pass

            match_to_display = False
            if mail_to:
                match_to_display = (agency_up in mail_to) or any(se in mail_to for se in search_emails)

            match_to = False
            if search_emails:
                try:
                    for i in range(1, mail.Recipients.Count + 1):
                        recip = mail.Recipients.Item(i)
                        addr = ""
                        name = ""
                        try:
                            addr = (recip.Address or "").upper()
                        except Exception:
                            pass
                        try:
                            name = (recip.Name or "").upper()
                        except Exception:
                            pass

                        for se in search_emails:
                            if se in addr or se in name or (mail_to and se in mail_to):
                                match_to = True
                                break
                        if match_to:
                            break
                except Exception:
                    pass

            return match_agency or match_to or match_to_display
        except Exception:
            return False

    def _find_original_mail(self, outlook, agency, recipients_to):
        """Busca el hilo original de la campaña en todos los almacenes/carpetas de Outlook."""
        try:
            mapi = outlook.GetNamespace("MAPI")
            
            # Buscar en todos los almacenes (Stores) de Outlook (incluye buzones compartidos)
            stores = mapi.Stores
            self._log(f"   [Debug] Buscando en {stores.Count} almacenes (Stores) de Outlook...")
            
            for s_idx in range(1, stores.Count + 1):
                try:
                    store = stores.Item(s_idx)
                    store_name = store.DisplayName
                    
                    # Intentar obtener carpetas por defecto de este almacén
                    sent_folder = None
                    inbox_folder = None
                    try:
                        sent_folder = store.GetDefaultFolder(5) # olFolderSentMail = 5
                    except:
                        pass
                    try:
                        inbox_folder = store.GetDefaultFolder(6) # olFolderInbox = 6
                    except:
                        pass
                    
                    folders = []
                    if sent_folder:
                        folders.append((f"{store_name} - Enviados", sent_folder, "[SentOn]"))
                    if inbox_folder:
                        folders.append((f"{store_name} - Entrada", inbox_folder, "[ReceivedTime]"))
                        
                    for folder_name, folder, sort_field in folders:
                        try:
                            items = folder.Items
                            try:
                                items.Sort(sort_field, True) # True para descendente (más recientes primero)
                            except Exception:
                                pass
                            
                            revisados = 0
                            for mail in items:
                                revisados += 1
                                if revisados > 500: # Limitar para evitar lentitud
                                    break
                                    
                                try:
                                    if self._mail_matches_original_thread(mail, agency, recipients_to):
                                        self._log(f"   (✓ Hilo original encontrado en {folder_name}: {mail.Subject})")
                                        return mail
                                except Exception:
                                    continue
                        except Exception as e:
                            logging.debug(f"Error al buscar en carpeta {folder_name}: {e}")
                except Exception as store_err:
                    logging.debug(f"Error procesando almacén: {store_err}")
                    
        except Exception as e:
            self._log(f"   (⚠ Error preparando búsqueda de hilo en almacenes: {e})")
            
        # Fallback a la carpeta por defecto de la sesión actual por si acaso
        try:
            self._log("   [Debug] Intentando fallback en carpetas de sesión por defecto...")
            folders = [
                ("Enviados por Defecto", mapi.GetDefaultFolder(5), "[SentOn]"),
                ("Entrada por Defecto", mapi.GetDefaultFolder(6), "[ReceivedTime]"),
            ]
            for folder_name, folder, sort_field in folders:
                try:
                    items = folder.Items
                    try:
                        items.Sort(sort_field, True)
                    except:
                        pass
                    revisados = 0
                    for mail in items:
                        revisados += 1
                        if revisados > 500:
                            break
                        try:
                            if self._mail_matches_original_thread(mail, agency, recipients_to):
                                return mail
                        except:
                            continue
                except:
                    pass
        except Exception as e:
            logging.debug(f"Error en fallback: {e}")
            
        return None

    def _save_complete_backup(self, agency, subject, recipients, html):
        """Guarda respaldo completo del correo principal: HTML + metadata JSON.
        
        Permite que cualquier usuario de la red pueda recuperar el correo original
        para construir recordatorios en hilo, aunque no haya sido el que lo envió.
        """
        try:
            agency_dir = self._get_agency_backup_dir(agency)
            
            # 1. Guardar HTML del cuerpo
            html_path = os.path.join(agency_dir, 'cuerpo_principal.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html)
            
            # 2. Guardar metadata JSON con toda la información de cabecera
            to_val = recipients.get('to', '') if isinstance(recipients, dict) else str(recipients)
            cc_val = recipients.get('cc', '') if isinstance(recipients, dict) else ''
            metadata = {
                'campaign': self.campaign,
                'agency': agency,
                'subject': subject,
                'to': to_val,
                'cc': cc_val,
                'sent_by': USER_NAME,
                'pc_name': PC_NAME,
                'saved_at': datetime.now().isoformat(),
                'excel_filename': os.path.basename(self.cronograma_path) if self.cronograma_path else ''
            }
            meta_path = os.path.join(agency_dir, 'metadata.json')
            with open(meta_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
            
            self._log(f"   (✓ Respaldo completo guardado para {agency} en: {agency_dir})")
            return agency_dir
        except Exception as be:
            self._log(f"   (⚠ No se pudo guardar el respaldo para {agency}: {be})")
            return None

    def _get_agency_backup_dir(self, agency):
        """Retorna (y crea si no existe) la subcarpeta de respaldo para una agencia específica."""
        safe_agency = agency.upper().replace(' ', '_').replace('/', '').replace('\\', '').replace(':', '')
        agency_dir = os.path.join(self._get_backup_dir(), safe_agency)
        os.makedirs(agency_dir, exist_ok=True)
        return agency_dir

    def _load_backup_metadata(self, agency):
        """Carga los metadatos y HTML de respaldo para la agencia indicada.
        
        Busca primero en la carpeta de la campaña actual, luego en toda la jerarquía
        de BACKUP_CORREOS de forma recursiva.
        
        Returns:
            dict con 'metadata', 'html', 'excel_path', o None si no existe nada.
        """
        safe_agency = agency.upper().replace(' ', '_').replace('/', '').replace('\\', '').replace(':', '')
        
        # 1. Buscar en la carpeta de la campaña actual
        agency_dir = os.path.join(self._get_backup_dir(), safe_agency)
        
        # 2. Fallback: buscar recursivamente en BACKUP_CORREOS
        if not os.path.isdir(agency_dir):
            try:
                if self.cronograma_path:
                    root_backup = os.path.join(os.path.dirname(os.path.abspath(self.cronograma_path)), 'BACKUP_CORREOS')
                else:
                    root_backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'BACKUP_CORREOS')
                if os.path.exists(root_backup):
                    for root, dirs, files in os.walk(root_backup):
                        if 'metadata.json' in files and os.path.basename(root) == safe_agency:
                            agency_dir = root
                            break
            except Exception:
                pass
        
        if not os.path.isdir(agency_dir):
            return None
        
        result = {}
        
        # Cargar metadata
        meta_path = os.path.join(agency_dir, 'metadata.json')
        if os.path.exists(meta_path):
            try:
                with open(meta_path, 'r', encoding='utf-8') as f:
                    result['metadata'] = json.load(f)
            except Exception:
                result['metadata'] = None
        else:
            result['metadata'] = None
        
        # Cargar HTML del cuerpo principal
        html_path = os.path.join(agency_dir, 'cuerpo_principal.html')
        if os.path.exists(html_path):
            try:
                with open(html_path, 'r', encoding='utf-8') as f:
                    result['html'] = f.read()
            except Exception:
                result['html'] = ''
        else:
            result['html'] = ''
        
        # Buscar Excel filtrado
        result['excel_path'] = None
        try:
            for fname in os.listdir(agency_dir):
                if fname.endswith('.xlsx'):
                    result['excel_path'] = os.path.join(agency_dir, fname)
                    break
        except Exception:
            pass
        
        return result if (result.get('metadata') or result.get('html')) else None

    def _get_backup_dir(self):
        """Devuelve la ruta para guardar/leer los backups organizados por Año y Campaña Bimensual.
        
        Prioridad:
          1. Ruta fija configurada por el usuario (backup_network_path)
          2. Ruta derivada de la ubicación del cronograma (fallback)
          3. Carpeta local ./data/BACKUP_CORREOS (fallback final)
        """
        now = datetime.now()
        year = str(now.year)
        
        # Determinar el nombre de la campaña bimensual
        if self.campaign and len(self.campaign.strip()) > 3:
            camp_name = self.campaign.strip()
            year_match = re.search(r'\b(20\d{2})\b', camp_name)
            if year_match:
                year = year_match.group(1)
        else:
            bimonthly_names = [
                'ENERO – FEBRERO', 'MARZO – ABRIL', 'MAYO – JUNIO',
                'JULIO – AGOSTO', 'SEPTIEMBRE – OCTUBRE', 'NOVIEMBRE – DICIEMBRE'
            ]
            idx = (now.month - 1) // 2
            camp_name = f"{bimonthly_names[idx]} {year}"
            
        folder_structure = os.path.join('BACKUP_CORREOS', year, camp_name)
        
        # 1. PRIORIDAD: Ruta fija configurada por el usuario
        if self.backup_network_path and os.path.isdir(self.backup_network_path):
            try:
                fixed_dir = os.path.join(self.backup_network_path, year, camp_name)
                os.makedirs(fixed_dir, exist_ok=True)
                return fixed_dir
            except Exception as e:
                logging.warning(f"No se pudo usar la ruta de backup configurada: {e}")
        
        # 2. FALLBACK: Derivar desde la carpeta del cronograma
        if self.cronograma_path:
            try:
                base_dir = os.path.dirname(os.path.abspath(self.cronograma_path))
                net_backup_dir = os.path.join(base_dir, folder_structure)
                os.makedirs(net_backup_dir, exist_ok=True)
                return net_backup_dir
            except Exception as e:
                logging.warning(f"No se pudo usar la carpeta de red para backups: {e}")
        
        # 3. FALLBACK FINAL: carpeta local
        local_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', folder_structure)
        os.makedirs(local_dir, exist_ok=True)
        return local_dir

    def _find_backup_file(self, agency):
        """Busca el archivo de respaldo de correo principal para la agencia en los últimos meses."""
        filename = f"principal_{agency.upper()}.html"
        
        # 1. Buscar en la carpeta de la campaña actual
        dir_now = self._get_backup_dir()
        path_now = os.path.join(dir_now, filename)
        if os.path.exists(path_now):
            return path_now
            
        # 2. Buscar en la carpeta de la campaña bimensual anterior (por si es cruzada)
        try:
            now = datetime.now()
            prev_date = now - timedelta(days=62)
            prev_year = str(prev_date.year)
            
            bimonthly_names = [
                'ENERO – FEBRERO',
                'MARZO – ABRIL',
                'MAYO – JUNIO',
                'JULIO – AGOSTO',
                'SEPTIEMBRE – OCTUBRE',
                'NOVIEMBRE – DICIEMBRE'
            ]
            idx = (prev_date.month - 1) // 2
            prev_camp_name = f"{bimonthly_names[idx]} {prev_year}"
            
            folder_structure = os.path.join('BACKUP_CORREOS', prev_year, prev_camp_name)
            
            if self.cronograma_path:
                base_dir = os.path.dirname(os.path.abspath(self.cronograma_path))
                dir_prev = os.path.join(base_dir, folder_structure)
            else:
                dir_prev = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', folder_structure)
                
            path_prev = os.path.join(dir_prev, filename)
            if os.path.exists(path_prev):
                return path_prev
        except Exception as e:
            logging.debug(f"Error buscando backup de campaña anterior: {e}")
            
        # 3. Fallback: buscar de forma recursiva en BACKUP_CORREOS por si acaso
        try:
            if self.cronograma_path:
                root_backup_dir = os.path.join(os.path.dirname(os.path.abspath(self.cronograma_path)), 'BACKUP_CORREOS')
            else:
                root_backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'BACKUP_CORREOS')
                
            if os.path.exists(root_backup_dir):
                for root, dirs, files in os.walk(root_backup_dir):
                    if filename in files:
                        return os.path.join(root, filename)
        except Exception as e:
            logging.debug(f"Error en búsqueda recursiva: {e}")
            
        return None

    def _get_signature_for_mail(self, mail):
        """Obtiene la firma del usuario de manera silenciosa o usando AppData."""
        try:
            initial_body = mail.HTMLBody
            if initial_body and len(initial_body.strip()) > 100:
                return initial_body
        except Exception:
            pass
            
        try:
            appdata = os.environ.get('APPDATA', '')
            sig_dir = os.path.join(appdata, 'Microsoft', 'Signatures')
            if os.path.exists(sig_dir):
                files = [f for f in os.listdir(sig_dir) if f.endswith('.htm') or f.endswith('.html')]
                if files:
                    files.sort(key=lambda x: os.path.getmtime(os.path.join(sig_dir, x)), reverse=True)
                    sig_path = os.path.join(sig_dir, files[0])
                    
                    for encoding in ('utf-16', 'utf-8', 'latin-1'):
                        try:
                            with open(sig_path, 'r', encoding=encoding, errors='ignore') as f:
                                content = f.read()
                                if content:
                                    return content
                        except Exception:
                            continue
        except Exception as e:
            logging.warning(f"Error cargando firma desde archivos: {e}")
            
        return ""

    def _attach_filtered_excel_to_mail(self, mail, agency):
        """Filtra el Excel de la campaña por agencia y lo adjunta al mail."""
        if self.cronograma_path:
            try:
                import tempfile as _tf, copy as _copy
                from openpyxl import load_workbook as _lw, Workbook as _WB

                wb_src = _lw(self.cronograma_path)
                ws_src = wb_src.active
                wb_out = _WB()
                ws_out = wb_out.active

                agencia_col_idx = None
                header_row_idx = None
                for row_idx, row in enumerate(ws_src.iter_rows(), start=1):
                    for cell in row:
                        if str(cell.value or '').strip().lower() in ['agencia', 'ag.', 'sede']:
                            agencia_col_idx = cell.column
                            header_row_idx = row_idx
                            break
                    if agencia_col_idx:
                        break

                def _copy_row(src_row, dst_ws, dst_row_idx):
                    for src_cell in src_row:
                        dst_cell = dst_ws.cell(row=dst_row_idx, column=src_cell.column, value=src_cell.value)
                        if src_cell.has_style:
                            dst_cell.font = _copy.copy(src_cell.font)
                            dst_cell.fill = _copy.copy(src_cell.fill)
                            dst_cell.border = _copy.copy(src_cell.border)
                            dst_cell.alignment = _copy.copy(src_cell.alignment)
                            dst_cell.number_format = src_cell.number_format

                out_row = 1
                for src_row_cells in ws_src.iter_rows():
                    row_idx = src_row_cells[0].row
                    if header_row_idx is None or row_idx <= header_row_idx:
                        _copy_row(src_row_cells, ws_out, out_row)
                        out_row += 1
                    else:
                        if agencia_col_idx:
                            cell_val = str(ws_src.cell(row=row_idx, column=agencia_col_idx).value or '').strip()
                            if cell_val.upper() == agency.upper():
                                _copy_row(src_row_cells, ws_out, out_row)
                                out_row += 1

                for col_letter, cd in ws_src.column_dimensions.items():
                    ws_out.column_dimensions[col_letter].width = cd.width

                tmp_xls_dir = _tf.mkdtemp(prefix='etv_gen_')
                orig_name = os.path.basename(self.cronograma_path)
                filtered_path = os.path.join(tmp_xls_dir, orig_name)
                
                wb_out.save(filtered_path)
                mail.Attachments.Add(os.path.abspath(filtered_path))
                self._log(f"   (📎 Excel filtrado adjunto: {orig_name})")
                
                # Guardar copia en la subcarpeta de agencia del respaldo compartido
                try:
                    agency_dir = self._get_agency_backup_dir(agency)
                    backup_xls_path = os.path.join(agency_dir, orig_name)
                    wb_out.save(backup_xls_path)
                    self._log(f"   (✓ Copia de Excel guardada en respaldo de agencia: {orig_name})")
                except Exception as b_err:
                    logging.debug(f"No se pudo guardar copia de Excel en respaldo: {b_err}")
                    
                return True
            except Exception as xe:
                try:
                    mail.Attachments.Add(os.path.abspath(self.cronograma_path))
                    self._log(f"   (⚠ Excel filtrado falló, adjuntando completo: {xe})")
                    return True
                except Exception as xe2:
                    self._log(f"   (✗ Error al adjuntar excel completo: {xe2})")
        return False

    def _on_auto_toggle(self):
        """Maneja la activación/desactivación del programador."""
        enabled = self.auto_enabled_var.get()
        if enabled:
            h_str = self.auto_hour_var.get().strip()
            if not re.match(r'^\d{2}:\d{2}$', h_str):
                messagebox.showerror("Formato de hora no válido", "Por favor ingresa la hora en formato HH:MM (ej. 08:30)")
                self.auto_enabled_var.set(False)
                return
            
            try:
                h, m = map(int, h_str.split(':'))
                if h < 0 or h > 23 or m < 0 or m > 59:
                    raise ValueError()
            except ValueError:
                messagebox.showerror("Hora no válida", "La hora debe estar entre 00:00 y 23:59")
                self.auto_enabled_var.set(False)
                return

            self._start_auto_scheduler()
            self._update_auto_status_ui(True)
            messagebox.showinfo(
                "Automatización Activada", 
                f"El servicio de automatización se ha iniciado.\n\nEl programa buscará recordatorios para mañana todos los días a las {h_str} y los enviará automáticamente."
            )
        else:
            self._stop_auto_scheduler()
            self._update_auto_status_ui(False)
            self._log("🤖 Automatización desactivada.")
            
        self._save_config()

    def _update_auto_status_ui(self, active):
        """Actualiza el texto descriptivo del estado de la automatización."""
        if active:
            h_str = self.auto_hour_var.get().strip()
            mode_str = "ENVIANDO automáticamente" if self.auto_send_var.get() else "generando BORRADORES"
            self.auto_status_label.config(
                text=f"Estado: ACTIVO — Esperando a las {h_str} para procesar recordatorios ({mode_str}). Mantenga el programa abierto.",
                fg='#27ae60'
            )
        else:
            self.auto_status_label.config(
                text="Estado: INACTIVO — No se enviará nada de forma programada.",
                fg='#636e72'
            )

    def _start_auto_scheduler(self):
        """Inicia el hilo del programador automático."""
        if hasattr(self, '_auto_thread_running') and self._auto_thread_running:
            return
            
        self._auto_thread_running = True
        self._auto_stop_event = threading.Event()
        
        thread = threading.Thread(target=self._auto_scheduler_loop, daemon=True)
        thread.start()
        
    def _stop_auto_scheduler(self):
        """Detiene el hilo del programador automático."""
        if hasattr(self, '_auto_stop_event'):
            self._auto_stop_event.set()
        self._auto_thread_running = False

    def _auto_scheduler_loop(self):
        self._log("🤖 Programador automático iniciado en segundo plano.")
        last_run_date = None
        
        while not self._auto_stop_event.is_set():
            if self._auto_stop_event.wait(15):
                break
                
            if not self.auto_enabled_var.get():
                break
                
            now = datetime.now()
            today = now.date()
            
            if last_run_date == today:
                continue
                
            scheduled_time_str = self.auto_hour_var.get().strip()
            try:
                sch_hour, sch_min = map(int, scheduled_time_str.split(':'))
            except Exception:
                sch_hour, sch_min = 8, 0
            
            # Usar ventana de ±90 segundos para no perder el trigger
            scheduled_dt = now.replace(hour=sch_hour, minute=sch_min, second=0, microsecond=0)
            diff_secs = abs((now - scheduled_dt).total_seconds())
            
            if diff_secs < 90:
                self._log("⏰ ¡Hora programada alcanzada! Iniciando envío de recordatorios automáticos...")
                try:
                    self._run_daily_automation()
                    last_run_date = today
                    self._log(f"✅ Automatización diaria finalizada para hoy. Próxima ejecución mañana a las {scheduled_time_str}.")
                except Exception as ex:
                    self._log(f"✗ ERROR crítico en automatización diaria: {ex}")

    def _run_daily_automation(self):
        """Ejecuta la automatización diaria de recordatorios para mañana."""
        if self.df_cronograma is None:
            self._log("⚠ Error automático: No hay cronograma cargado. Cargue el cronograma primero.")
            return
            
        tomorrow = date.today() + timedelta(days=1)
        # Viernes → buscar para el lunes; Sábado → buscar para el lunes
        if tomorrow.weekday() == 5:   # Sábado
            tomorrow = tomorrow + timedelta(days=2)
        elif tomorrow.weekday() == 6: # Domingo
            tomorrow = tomorrow + timedelta(days=1)
        tomorrow_str = tomorrow.strftime('%d/%m/%Y')
        
        df = self.df_cronograma.copy()
        if 'Target_Date_Calc' in df.columns:
            df_tomorrow = df[df['Target_Date_Calc'] == tomorrow]
        else:
            date_col = 'Fecha_Entrega' if 'Fecha_Entrega' in df.columns else 'Fecha_Pago'
            df['Target_Date'] = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
            df_tomorrow = df[df['Target_Date'].dt.date == tomorrow]
            
        agencies = sorted(df_tomorrow['Agencia'].unique().tolist()) if not df_tomorrow.empty else []
        
        if not agencies:
            self._log(f"🤖 Automatización: Sin pagos programados para el {tomorrow_str}. No se enviará nada.")
            return
            
        self._log(f"🤖 Automatización: Se detectaron {len(agencies)} agencias para mañana {tomorrow_str}: {agencies}")
        
        import pythoncom
        pythoncom.CoInitialize()
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            
            results = []
            auto_send = self.auto_send_var.get()
            for i, agency in enumerate(agencies):
                self._log(f"🤖 Procesando [{i+1}/{len(agencies)}] {agency}...")
                try:
                    result = self._create_reminder_draft(outlook, agency, auto_send=auto_send)
                    results.append(result)
                except Exception as inner_ex:
                    self._log(f"   ✗ Error: {inner_ex}")
                    results.append({'agency': agency, 'status': 'ERROR', 'to': '', 'error': str(inner_ex)})
                    
            self._report_to_portal(results)
        finally:
            pythoncom.CoUninitialize()
    


    def _save_test_html(self, agency, subject, html_body):
        """Genera y guarda el HTML localmente para probar sin Outlook."""
        import tempfile, webbrowser
        tmp_dir = tempfile.gettempdir()
        safe_ag = str(agency).replace(' ', '_').replace('/', '').replace('\\', '')
        filepath = os.path.join(tmp_dir, f"Prueba_Correos_ETV_{safe_ag}.html")
        
        full_html = f"<html><head><meta charset='utf-8'><title>{subject}</title></head><body style='padding:20px;'><h2 style='color:#C8102E;'>ASUNTO: {subject}</h2><hr><br>{html_body}</body></html>"
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(full_html)
            
        if not hasattr(self, '_test_html_opened'):
            webbrowser.open('file://' + os.path.abspath(filepath))
            self._test_html_opened = True
            
        self._log(f"   (📎 HTML local generado: {filepath})")
        return {'agency': agency, 'status': 'GENERADO', 'to': 'MODO_PRUEBA_LOCAL'}

    def _get_recipients(self, agency):
        """Obtiene destinatarios de una agencia.

        Si el dataframe de agencias no tiene la columna esperada o está ausente,
        devolvemos cadenas vacías en lugar de fallar con KeyError.
        """
        if self.df_agencias is None:
            self._log("   [Debug] df_agencias es None (no se cargó el archivo de correos)")
            return {'to': '', 'cc': ''}

        if 'Agencia' not in self.df_agencias.columns:
            self._log(f"   [Debug] Columna 'Agencia' no encontrada en el excel de correos. Columnas: {list(self.df_agencias.columns)}")
            return {'to': '', 'cc': ''}

        try:
            # Normalizar nombre de agencia para búsqueda
            ag_upper = str(agency).strip().upper()
            self._log(f"   [Debug] Buscando destinatarios para: '{ag_upper}'")
            
            row = self.df_agencias[self.df_agencias['Agencia'].astype(str).str.strip().str.upper() == ag_upper]
        except Exception as e:
            self._log(f"   [Debug] Error al buscar en df_agencias: {e}")
            return {'to': '', 'cc': ''}

        if not row.empty:
            dest_to = str(row.iloc[0].get('Correo', ''))
            dest_cc = str(row.iloc[0].get('CC', '')) if 'CC' in row.columns and pd.notna(row.iloc[0].get('CC', '')) else ''
            self._log(f"   [Debug] Encontrado: To={dest_to}, CC={dest_cc}")
            return {
                'to': dest_to,
                'cc': dest_cc
            }
        
        self._log(f"   [Debug] No se encontró la agencia '{ag_upper}' en agencias_correos.xlsx")
        # Opcional: mostrar las primeras 5 agencias para depurar
        known_ags = self.df_agencias['Agencia'].astype(str).str.strip().str.upper().unique()[:5]
        self._log(f"   [Debug] Algunas agencias conocidas: {list(known_ags)}...")
        return {'to': '', 'cc': ''}

    def _build_html_body(self, agency, fecha, df, has_juntos, has_contigo):
        """Construye el cuerpo HTML del correo con tono profesional."""
        programas = ["PENSIÓN 65"]
        if has_juntos: programas.append("JUNTOS")
        if has_contigo: programas.append("CONTIGO")
        
        if len(programas) > 1:
            partes_bold = [f'<b>{p}</b>' for p in programas]
            prog_mencion = "los Programas Sociales " + ", ".join(partes_bold[:-1]) + " y " + partes_bold[-1]
        else:
            prog_mencion = "el Programa Social <b>" + programas[0] + "</b>"

        table = self._build_html_table(df, has_juntos, has_contigo)
        
        return f'''<div style="font-family:Calibri,sans-serif;font-size:11pt;color:#222;">
<p>Estimados buenos días,<br><b>BN &ndash; {agency}</b></p>
<p>Por encargo del Sr. <b>Hermes Jonathan Bazán Méndez - Jefe de Caja y Valores</b>, y por medio de la presente, se solicita programar la entrega de fondos, que inicia el <b><span style="background:yellow">{fecha}</span></b>, con el fin de realizar los pagos a los usuarios de {prog_mencion}, les recordamos entregar los fondos a las ETV correspondientes a la campaña de pagos {self.campaign}.</p>
<p><b><span style="font-size:12pt;color:#C00000">Se adjunta archivo en formato Excel con la programación detallada de la presente campaña.</span></b></p>
<p><i><b><span style="font-size:11pt;background:aqua">Obs.: En el caso de la ETV PROSEGUR, se deberá imprimir y entregar las planillas, usando el aplicativo PSJN.</span></b></i></p>
<p>Agradecemos de antemano su atención y coordinación para el cumplimiento oportuno de estas obligaciones.</p>
<br>{table}<br></div>'''

    def _build_html_table(self, df, has_juntos, has_contigo=False, is_reminder=False):
        """Construye tabla HTML con formato y colores similares al Excel original."""
        columns = []
        
        # Grupo: Departamento/Provincia/... (Salmón claro)
        grp1_bg = '#FF9999'
        columns.append({'group': 'Departamento', 'gbg': grp1_bg, 'cbg': grp1_bg, 'col': 'Departamento', 'cl': 'black'})
        columns.append({'group': 'Provincia', 'gbg': grp1_bg, 'cbg': grp1_bg, 'col': 'Provincia', 'cl': 'black'})
        if 'Cod_Pueblo' in df.columns:
            columns.append({'group': 'Cod.<br>Pueblo', 'gbg': grp1_bg, 'cbg': grp1_bg, 'col': 'Cod_Pueblo', 'cl': 'black', 'is_code': True})
        columns.append({'group': 'Distrito', 'gbg': grp1_bg, 'cbg': grp1_bg, 'col': 'Distrito', 'cl': 'black'})
        if 'Cod_Pto_Pago' in df.columns:
            columns.append({'group': 'Cod.<br>Pto.Pago', 'gbg': grp1_bg, 'cbg': grp1_bg, 'col': 'Cod_Pto_Pago', 'cl': 'black', 'is_code': True})
            
        # Grupo: Punto de Pago (Rojo)
        columns.append({'group': 'Punto de Pago', 'gbg': '#FF0000', 'cbg': '#FF0000', 'col': 'Punto de Pago', 'cl': 'black'})
        
        # Grupo: Pensión 65 (Amarillo)
        grp3_bg = '#FFFF99'
        columns.append({'group': 'Pensión 65', 'gbg': grp3_bg, 'cbg': grp3_bg, 'col': 'N° Usuarios', 'cl': 'black'})
        columns.append({'group': 'Pensión 65', 'gbg': grp3_bg, 'cbg': grp3_bg, 'col': 'Importe', 'cl': 'black'})
        
        # Grupo: Juntos (Verde)
        if has_juntos:
            grp4_bg = '#00B050'
            columns.append({'group': 'Juntos', 'gbg': grp4_bg, 'cbg': grp4_bg, 'col': 'N° Usuarios', 'cl': 'black'})
            columns.append({'group': 'Juntos', 'gbg': grp4_bg, 'cbg': grp4_bg, 'col': 'Importe', 'cl': 'black'})
            
        # Grupo: Contigo (Celeste)
        if has_contigo:
            grp5_bg = '#00B0F0'
            columns.append({'group': 'Contigo', 'gbg': grp5_bg, 'cbg': grp5_bg, 'col': 'N° Usuarios', 'cl': 'black'})
            columns.append({'group': 'Contigo', 'gbg': grp5_bg, 'cbg': grp5_bg, 'col': 'Importe', 'cl': 'black'})
            
        # Grupo: Fecha y Hora de Pago (Verde agua)
        grp6_bg = '#CCFFCC'
        columns.append({'group': 'Fecha y Hora de<br>Pago PP65', 'gbg': grp6_bg, 'cbg': grp6_bg, 'col': 'Fecha Pago', 'cl': 'black'})
        if 'Hora_Pago' in df.columns:
            columns.append({'group': 'Fecha y Hora de<br>Pago PP65', 'gbg': grp6_bg, 'cbg': grp6_bg, 'col': 'Hora', 'cl': 'black'})
            
        # Grupo: Empresa Responsable (Verde oscuro)
        if 'ETV' in df.columns:
            columns.append({'group': 'Empresa<br>Responsable', 'gbg': '#00B050', 'cbg': '#00B050', 'col': 'ETV', 'cl': 'black'})
            
        # Grupo: Preparación y Entrega (Celeste claro / Azul)
        grp8_bg = '#00B0F0'
        columns.append({'group': 'Preparación y ENTREGA DE FONDOS', 'gbg': grp8_bg, 'cbg': grp8_bg, 'col': 'Agencia', 'cl': 'black'})
        if 'Fecha_Entrega' in df.columns:
            columns.append({'group': 'Preparación y ENTREGA DE FONDOS', 'gbg': grp8_bg, 'cbg': grp8_bg, 'col': 'Fecha', 'cl': 'black', 'hl': True})
        if 'Hora_Entrega' in df.columns:
            columns.append({'group': 'Preparación y ENTREGA DE FONDOS', 'gbg': grp8_bg, 'cbg': grp8_bg, 'col': 'Hora', 'cl': 'black', 'hl': True})

        # Construir tabla (2 filas de cabecera con rowspan inteligente)
        html = '<table style="border-collapse:collapse;font-family:Calibri;width:100%;"><thead><tr>'

        # Agrupar columnas manteniendo orden de aparicion
        seen_groups = {}
        group_order = []
        for c in columns:
            g = c['group']
            if g not in seen_groups:
                seen_groups[g] = []
                group_order.append(g)
            seen_groups[g].append(c)
        groups = [(g, seen_groups[g]) for g in group_order]

        th_style = 'padding:0px 2px;border:1px solid #ccc;text-align:center;font-weight:bold;font-size:8pt;vertical-align:middle;'

        # Fila 1: Grupos (con rowspan=2 si el grupo tiene una sola columna)
        row2_cells = []
        for grp_name, grp_cols in groups:
            grp_bg = grp_cols[0]['gbg']
            grp_span = len(grp_cols)
            if grp_span == 1:
                # Celda unica: rowspan=2 para no duplicar texto en fila 2
                html += f'<td rowspan="2" style="background:{grp_bg};color:black;{th_style}">{grp_name}</td>'
            else:
                html += f'<td colspan="{grp_span}" style="background:{grp_bg};color:black;{th_style}">{grp_name}</td>'
                row2_cells.extend(grp_cols)
        html += '</tr><tr>'

        # Fila 2: Solo sub-columnas de grupos multi-columna
        for col in row2_cells:
            html += f'<td style="background:{col["cbg"]};color:black;{th_style}">{col["col"]}</td>'
        html += '</tr></thead><tbody>'
        
        td_s = 'padding:0px 2px;border:1px solid #ccc;text-align:center;font-size:8pt;'
        for i, (_, row) in enumerate(df.iterrows()):
            bg = 'background:#f9f0f0;' if (not is_reminder and i % 2 == 1) else ''
            html += f'<tr style="{bg}">'
            
            for col in columns:
                # Mapear nombre de columna a clave en el dataframe
                group_key = col['group'].replace('Pensión 65', 'P65')
                mapping = {
                    'Punto de Pago': 'Punto_de_Pago',
                    'N° Usuarios': group_key + '_Usuarios',
                    'Importe': group_key + '_Importe',
                    'Fecha Pago': 'Fecha_Pago',
                    'Hora': 'Hora_Pago' if col['group'].startswith('Fecha') else 'Hora_Entrega',
                    'Fecha': 'Fecha_Entrega',
                    'Agencia': 'Agencia',
                    'ETV': 'ETV',
                    'Distrito': 'Distrito',
                    'Provincia': 'Provincia',
                    'Departamento': 'Departamento',
                    'Cod_Pueblo': 'Cod_Pueblo',
                    'Cod_Pto_Pago': 'Cod_Pto_Pago'
                }
                actual_col = mapping.get(col['col'], col['col'])
                v = row.get(actual_col, '')
                
                # Formatear valor
                is_code = col.get('is_code', False)
                if is_code:
                    # Preservar ceros a la izquierda (ej: 664 -> 0664, 85 -> 0085)
                    if pd.notna(v) and str(v).strip() not in ('', 'nan', 'None'):
                        raw = str(v).strip()
                        try:
                            # Convertir float->int para eliminar el decimal, luego rellenar a 4 cifras
                            num = int(float(raw))
                            v_str = str(num).zfill(4)
                        except (ValueError, TypeError):
                            # Si ya es un string con ceros (ej: '0664' literal), mantener tal cual
                            v_str = raw
                    else:
                        v_str = ''
                elif 'Importe' in col['col']:
                    v_str = self._fmt_amt(v)
                elif 'Usuarios' in col['col']:
                    v_str = self._fmt_int(v)
                elif 'Fecha' in col['col']:
                    v_str = self._fmt_date(v)
                elif 'Hora' in col['col']:
                    v_str = self._fmt_time(v)
                else:
                    v_str = str(v).strip() if pd.notna(v) else ''

                # Aplicar estilos: amarillo en Importe (siempre) y en celdas hl
                style = td_s
                if 'Importe' in col['col']:
                    if v_str:
                        style += 'background:#FFFF00;font-weight:bold;'
                    else:
                        # Importe vacío = rojo
                        html += f'<td style="{style}background:red;"></td>'
                        continue
                elif col.get('hl'):
                    style += 'background:#FFFF00;font-weight:bold;'

                if not v_str:
                    html += f'<td style="{style}background:red;"></td>'
                else:
                    html += f'<td style="{style}">{v_str}</td>'
            html += '</tr>'

        html += '</tbody></table>'
        return html

    @staticmethod
    def _fmt_date(v):
        try:
            return pd.to_datetime(v).strftime('%d/%m/%y')
        except:
            return str(v) if pd.notna(v) else ''
    
    @staticmethod
    def _fmt_time(v):
        try:
            if hasattr(v, 'strftime'):
                return v.strftime('%I:%M %p')
            return str(v)
        except:
            return str(v) if pd.notna(v) else ''
    
    @staticmethod
    def _fmt_amt(v):
        try:
            val = float(v)
            return f"{val:,.2f}" if pd.notna(v) and val != 0 else ''
        except:
            return ''
    
    @staticmethod
    def _fmt_int(v):
        try:
            val = float(v)
            return str(int(val)) if pd.notna(v) and val != 0 else ''
        except:
            return ''

    def _report_to_portal(self, results):
        """Envía el reporte al portal central."""
        url = self.server_entry.get().strip()
        if not url:
            self._log("\nReporte al portal: Deshabilitado (no hay URL configurada)")
            return

        try:
            import urllib.request
            data = json.dumps({
                'user_name': USER_NAME,
                'user_pc': PC_NAME,
                'campaign': self.campaign,
                'campaign_code': self.campaign_code,
                'results': results
            }).encode('utf-8')

            req = urllib.request.Request(
                f"{url}/api/report",
                data=data,
                headers={'Content-Type': 'application/json'}
            )
            urllib.request.urlopen(req, timeout=5)
            self._log(f"\n✓ Reporte enviado al portal: {len(results)} registros")
        except Exception as e:
            self._log(f"\n⚠ No se pudo reportar al portal: {e}")

    def _log(self, message):
        """Agrega mensaje al log."""
        timestamp = datetime.now().strftime('%H:%M:%S')
        full_msg = f"[{timestamp}] {message}\n"
        
        self.root.after(0, lambda: self.log_text.insert(tk.END, full_msg))
        self.root.after(0, lambda: self.log_text.see(tk.END))
        
        logging.info(message)

    def run(self):
        """Inicia la aplicación."""
        self.root.mainloop()


# ─── Entry Point ──────────────────────────────────────

if __name__ == '__main__':
    try:
        app = CorreosETVApp()
        app.run()
    except Exception as e:
        logging.error(f"Error fatal: {e}", exc_info=True)
        messagebox.showerror("Error Fatal", f"La aplicación ha encontrado un error crítico:\n\n{e}")
