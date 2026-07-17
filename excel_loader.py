"""
excel_loader.py
Lectura de archivos Excel (.xlsx) y emparejamiento automático de sus
columnas con los campos {campo} detectados en la plantilla Word.
"""

import re
import unicodedata
from openpyxl import load_workbook


def _normalize(text):
    """Normaliza texto para comparar nombres de columnas/campos sin
    importar tildes, mayúsculas, espacios o guiones."""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def read_excel(excel_path, sheet_name=None):
    """Lee la primera fila como encabezados y devuelve una tupla:
    (columnas_originales, lista_de_filas_como_diccionario)."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in headers_raw]

    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row_dict = {}
        for header, value in zip(headers, raw_row):
            if not header:
                continue
            row_dict[header] = "" if value is None else value
        rows.append(row_dict)

    return headers, rows


def _token_overlap_score(norm_a, norm_b):
    """Calcula qué tan parecidos son dos nombres normalizados comparando
    sus palabras (tokens) separadas por '_'. Ignora palabras de relleno
    comunes en español ('de', 'del', 'la', etc). Devuelve un score entre
    0 y 1."""
    stopwords = {"de", "del", "la", "el", "los", "las", "y"}
    tokens_a = set(norm_a.split("_")) - stopwords
    tokens_b = set(norm_b.split("_")) - stopwords
    if not tokens_a or not tokens_b:
        return 0.0
    overlap = tokens_a & tokens_b
    return len(overlap) / max(len(tokens_a), len(tokens_b))


def auto_map_columns(excel_columns, template_fields):
    """Intenta emparejar cada campo de la plantilla con una columna del
    Excel comparando nombres normalizados: primero coincidencia exacta,
    luego alias conocidos, substring, y por último similitud de palabras.
    Devuelve un diccionario {campo_plantilla: columna_excel_o_None}."""
    norm_columns = {_normalize(c): c for c in excel_columns if c}

    # Definir alias comunes para campos de correspondencia en el banco
    aliases = {
        "cuenta": ["numero_de_cuenta", "nro_de_cuenta", "nro_cuenta", "cuenta_bancaria", "cuenta"],
        "cliente": ["razon_social", "nombre", "destinatario", "nombre_cliente", "cliente"],
        "direccion": ["domicilio", "direccion_domicilio", "direccion_de_domicilio", "direccion"],
        "destinatario_correo": ["direccion_de_correo", "correo", "email", "mail", "correo_destinatario", "direccion_de_correo_electronico"],
        "copia_correo": ["cc", "copia", "con_copia"],
        "celular": ["numero_de_celular", "nro_celular", "telefono", "celular"],
    }

    mapping = {}
    for field in template_fields:
        norm_field = _normalize(field)
        match = None

        # 1. Coincidencia exacta
        if norm_field in norm_columns:
            mapping[field] = norm_columns[norm_field]
            continue

        # 2. Búsqueda por alias semánticos conocidos
        if norm_field in aliases:
            for alias_key in aliases[norm_field]:
                if alias_key in norm_columns:
                    match = norm_columns[alias_key]
                    break

        # 3. Búsqueda por substring
        if not match:
            for norm_col, original_col in norm_columns.items():
                if norm_field in norm_col or norm_col in norm_field:
                    match = original_col
                    break

        # 4. Búsqueda por solapamiento de palabras (Jaccard-like score)
        if not match:
            best_score = 0.0
            for norm_col, original_col in norm_columns.items():
                score = _token_overlap_score(norm_field, norm_col)
                if score > best_score:
                    best_score = score
                    match = original_col
            if best_score < 0.5:
                match = None

        mapping[field] = match

    return mapping


def build_rows_for_fields(template_fields, excel_columns, excel_rows, mapping):
    """Convierte las filas del Excel (indexadas por nombre de columna
    original) en filas indexadas por nombre de campo de la plantilla,
    usando el diccionario `mapping` ({campo: columna})."""
    result = []
    for excel_row in excel_rows:
        new_row = {}
        for field in template_fields:
            col = mapping.get(field)
            new_row[field] = excel_row.get(col, "") if col else ""
        result.append(new_row)
    return result
